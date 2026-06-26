"""
app.py — Monitor Trato Directo SSMOCC v10
Navegación sin sidebar para evitar problema de colapso en Streamlit Cloud
"""
import streamlit as st
import hashlib, json, datetime, io, base64, csv
from pathlib import Path
import urllib.request, urllib.error

st.set_page_config(
    page_title="Monitor TD — SSMOCC",
    page_icon="🏥",
    layout="wide",
    initial_sidebar_state="collapsed",
)

st.markdown("""
<style>
#MainMenu {visibility:hidden;}
footer {visibility:hidden;}
header {visibility:hidden;}
[data-testid="stSidebarNavItems"]{display:none!important;}
[data-testid="stSidebarNav"]{display:none!important;}
[data-testid="collapsedControl"]{display:none!important;}
section[data-testid="stSidebar"]{display:none!important;}
.block-container{padding-top:0!important;padding-bottom:2rem!important;max-width:100%!important;}
/* Navbar top */
.topbar{background:#0A3964;padding:0;margin-bottom:0;position:sticky;top:0;z-index:999;}
.topbar-inner{display:flex;align-items:center;gap:0;padding:0 8px;}
.topbar-logo{padding:8px 12px;border-right:3px solid #C0392B;display:flex;align-items:center;}
.topbar-logo img{height:38px;border-radius:3px;}
.topbar-title{padding:0 16px;flex:1;}
.topbar-title .t1{font-size:14px;font-weight:700;color:#fff;}
.topbar-title .t2{font-size:10px;color:rgba(255,255,255,.5);}
.topbar-user{padding:8px 14px;border-left:1px solid rgba(255,255,255,.15);text-align:right;}
.topbar-user .u1{font-size:12px;font-weight:600;color:#fff;}
.topbar-user .u2{font-size:10px;color:rgba(255,255,255,.45);}
/* Nav pills */
.navpills{background:#1F3864;display:flex;gap:4px;padding:6px 16px;flex-wrap:wrap;align-items:center;}
.stButton>button{border-radius:6px!important;font-size:12px!important;padding:5px 14px!important;}
/* Cards */
.kcard{background:white;border:0.5px solid #e2e8f0;border-radius:10px;padding:14px 16px;border-top:3px solid #1F3864;}
.kcard .lbl{font-size:10px;text-transform:uppercase;letter-spacing:.06em;color:#64748b;margin-bottom:5px;}
.kcard .val{font-size:22px;font-weight:700;line-height:1;}
.kcard .note{font-size:11px;color:#94a3b8;margin-top:4px;}
/* Niveles */
.nr{background:#FEF2F2;color:#991B1B;border:1px solid #FECACA;padding:2px 10px;border-radius:5px;font-size:11px;font-weight:700;}
.na{background:#FFFBEB;color:#92400E;border:1px solid #FDE68A;padding:2px 10px;border-radius:5px;font-size:11px;font-weight:700;}
.nv{background:#F0FDF4;color:#166534;border:1px solid #BBF7D0;padding:2px 10px;border-radius:5px;font-size:11px;font-weight:700;}
</style>
""", unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════════════
# DATOS
# ═══════════════════════════════════════════════════════════════════════
# ── ALMACENAMIENTO PERSISTENTE ──────────────────────────────────────────
# Usa GitHub como backend para persistencia real en Streamlit Cloud
# Configura st.secrets con:
#   [github]
#   token = "ghp_..."
#   repo  = "barolaro/ssmoc_td"
#   branch = "main"

DATA_DIR = Path(__file__).parent / "data"
DATA_DIR.mkdir(exist_ok=True)
USERS_FILE   = DATA_DIR / "users.json"
REPORTS_FILE = DATA_DIR / "reports.json"
DATOS_FILE   = DATA_DIR / "datos_minsal.json"
DATOS_PERIODOS_FILE = DATA_DIR / "datos_periodos.json"

def _gh_cfg():
    """Retorna config de GitHub desde secrets si está disponible."""
    try:
        s = st.secrets.get("github", {})
        if s.get("token") and s.get("repo"):
            return s
    except Exception:
        pass
    return None

def _gh_read(filename: str):
    """Lee un archivo JSON desde GitHub. Nunca lanza excepción."""
    cfg = _gh_cfg()
    if not cfg:
        return None, ""
    url = f"https://api.github.com/repos/{cfg['repo']}/contents/data/{filename}"
    req = urllib.request.Request(url, headers={
        "Authorization": f"token {cfg['token']}",
        "Accept": "application/vnd.github.v3+json"
    })
    try:
        with urllib.request.urlopen(req, timeout=10) as r:
            data = json.loads(r.read())
            content_b64 = data.get("content","").replace("\n","")
            return json.loads(base64.b64decode(content_b64).decode("utf-8")), data.get("sha","")
    except Exception:
        # 404 = archivo no existe aún, otros errores = red o token
        return None, ""

def _gh_write(filename: str, content_obj, sha: str = ""):
    """Escribe un archivo JSON en GitHub."""
    cfg = _gh_cfg()
    if not cfg: return False
    url = f"https://api.github.com/repos/{cfg['repo']}/contents/data/{filename}"
    content_bytes = json.dumps(content_obj, ensure_ascii=False, indent=2, default=str).encode("utf-8")
    payload = {
        "message": f"Update {filename} via SSMOCC platform",
        "content": base64.b64encode(content_bytes).decode("utf-8"),
        "branch": cfg.get("branch", "main"),
    }
    if sha: payload["sha"] = sha
    req = urllib.request.Request(url,
        data=json.dumps(payload).encode("utf-8"),
        headers={"Authorization": f"token {cfg['token']}",
                 "Accept": "application/vnd.github.v3+json",
                 "Content-Type": "application/json"},
        method="PUT")
    try:
        with urllib.request.urlopen(req, timeout=15):
            return True
    except Exception:
        return False

def _load_json_persistent(local_path: Path, filename: str, default):
    """Carga JSON: primero GitHub, luego local, luego default. Nunca lanza excepción."""
    try:
        result = _gh_read(filename)
        if result and result[0] is not None:
            data, sha = result
            try:
                local_path.write_text(json.dumps(data, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
            except Exception:
                pass
            return data
    except Exception:
        pass
    # Try local cache
    try:
        if local_path.exists():
            return json.loads(local_path.read_text(encoding="utf-8"))
    except Exception:
        pass
    return default

def _save_json_persistent(local_path: Path, filename: str, data):
    """Guarda JSON: local + GitHub. Nunca lanza excepción."""
    try:
        txt = json.dumps(data, ensure_ascii=False, indent=2, default=str)
        local_path.write_text(txt, encoding="utf-8")
    except Exception:
        pass
    try:
        result = _gh_read(filename)
        sha = result[1] if result and result[1] else ""
        _gh_write(filename, data, sha)
    except Exception:
        pass

def _hash(p): return hashlib.sha256(p.encode()).hexdigest()

ESTABLECIMIENTOS = {
    "traumatologico": {"nombre":"Instituto Traumatológico Dr. Teodoro Gebauer","nombre_corto":"Inst. Traumatológico","rut":"61.608.203-5","codigo_deis":"110110","pct_2026":62.17,"pct_2025":37.78,"brecha":46.17,"variacion":24.39,"nivel":"rojo","denominador":5868063249,"numerador":3648430684},
    "direccion":      {"nombre":"Dirección del Servicio Metropolitano Occidente","nombre_corto":"Dir. SSMOCC","rut":"61.608.200-0","codigo_deis":"110010","pct_2026":37.16,"pct_2025":5.30,"brecha":21.16,"variacion":31.86,"nivel":"rojo","denominador":5835193133,"numerador":2168191360},
    "felix_bulnes":   {"nombre":"Hospital Dr. Félix Bulnes Cerda","nombre_corto":"H. Félix Bulnes","rut":"61.608.205-1","codigo_deis":"110120","pct_2026":26.44,"pct_2025":21.59,"brecha":10.44,"variacion":4.85,"nivel":"rojo","denominador":19487598854,"numerador":5152786649},
    "san_juan":       {"nombre":"Hospital San Juan de Dios","nombre_corto":"H. San Juan de Dios","rut":"61.608.204-3","codigo_deis":"110100","pct_2026":11.61,"pct_2025":10.70,"brecha":-4.39,"variacion":0.91,"nivel":"amarillo","denominador":39066763291,"numerador":4535664841},
    "crs_allende":    {"nombre":"Centro de Referencia Salud Occidente Salvador Allende","nombre_corto":"CRS Salvador Allende","rut":"61.933.400-0","codigo_deis":"110300","pct_2026":7.07,"pct_2025":4.76,"brecha":-8.93,"variacion":2.31,"nivel":"amarillo","denominador":1954204231,"numerador":138204661},
    "melipilla":      {"nombre":"Hospital San José (Melipilla)","nombre_corto":"H. Melipilla","rut":"61.602.122-2","codigo_deis":"110150","pct_2026":6.60,"pct_2025":8.29,"brecha":-9.40,"variacion":-1.69,"nivel":"verde","denominador":6908126647,"numerador":455863218},
    "penaflor":       {"nombre":"Hospital de Peñaflor","nombre_corto":"H. Peñaflor","rut":"61.602.121-4","codigo_deis":"110140","pct_2026":6.89,"pct_2025":17.58,"brecha":-9.11,"variacion":-10.69,"nivel":"verde","denominador":1677287316,"numerador":115598031},
    "curacavi":       {"nombre":"Hospital de Curacaví","nombre_corto":"H. Curacaví","rut":"61.602.125-7","codigo_deis":"110160","pct_2026":5.62,"pct_2025":14.42,"brecha":-10.38,"variacion":-8.80,"nivel":"verde","denominador":1001068453,"numerador":56257312},
    "talagante":      {"nombre":"Hospital Adalberto Steeger (Talagante)","nombre_corto":"H. Talagante","rut":"61.602.121-4","codigo_deis":"110130","pct_2026":3.21,"pct_2025":10.63,"brecha":-12.79,"variacion":-7.42,"nivel":"verde","denominador":7242041556,"numerador":232416794},
}

# Copia base para poder cambiar de período sin arrastrar datos del período anterior.
BASE_ESTABLECIMIENTOS = json.loads(json.dumps(ESTABLECIMIENTOS))

# Mapa de DEIS → clave interna (para actualización desde CSV MINSAL)
DEIS_MAP = {
    "110110": "traumatologico",
    "110010": "direccion",
    "110120": "felix_bulnes",
    "110100": "san_juan",
    "110300": "crs_allende",
    "110150": "melipilla",
    "110140": "penaflor",
    "110160": "curacavi",
    "110130": "talagante",
}

def load_datos_minsal():
    """Carga datos MINSAL desde GitHub/local. Nunca se pierden."""
    saved = _load_json_persistent(DATOS_FILE, "datos_minsal.json", {})
    if saved:
        for eid, vals in saved.items():
            if eid in ESTABLECIMIENTOS:
                ESTABLECIMIENTOS[eid].update(vals)

def save_datos_minsal(updates: dict):
    """Guarda datos MINSAL en GitHub + local. Persiste entre reinicios."""
    _save_json_persistent(DATOS_FILE, "datos_minsal.json", updates)

def load_datos_periodos():
    """Carga la base histórica: año → período → establecimientos."""
    return _load_json_persistent(DATOS_PERIODOS_FILE, "datos_periodos.json", {})

def save_datos_periodo(year: int, reporte_id: str, updates: dict, fuente: str = "CSV MINSAL"):
    """Guarda una carga oficial MINSAL asociada a un año y período específico."""
    data = load_datos_periodos()
    y = str(year)
    data.setdefault(y, {})
    data[y][reporte_id] = {"metadata": {"year": year, "reporte_id": reporte_id, "fuente": fuente, "fecha_carga": datetime.datetime.now().isoformat(timespec="seconds"), "establecimientos": len(updates)}, "establecimientos": updates}
    _save_json_persistent(DATOS_PERIODOS_FILE, "datos_periodos.json", data)

def get_periodo_info(reporte_id: str):
    return next((p for p in PERIODOS if p["id"] == reporte_id), PERIODOS[0])

def periodo_display(year: int, reporte_id: str):
    p = get_periodo_info(reporte_id)
    return f"{year} · {p['label']} · {p['periodo']}"

def periodo_tiene_datos(year: int = None, reporte_id: str = None) -> bool:
    """Indica si el año/período seleccionado tiene una carga oficial MINSAL registrada."""
    year = year or st.session_state.get("selected_year", 2026)
    reporte_id = reporte_id or st.session_state.get("selected_report", "R1")
    data = load_datos_periodos()
    return bool(data.get(str(year), {}).get(reporte_id, {}).get("establecimientos"))

def mensaje_periodo_sin_datos(year: int = None, reporte_id: str = None, detalle: bool = True):
    """Mensaje institucional cuando el período aún no tiene carga oficial."""
    year = year or st.session_state.get("selected_year", 2026)
    reporte_id = reporte_id or st.session_state.get("selected_report", "R1")
    pinfo = get_periodo_info(reporte_id)
    texto_detalle = ""
    if detalle:
        texto_detalle = "<br>Para evitar distorsiones, no se muestran indicadores, semáforos, reportes, boletines ni exportaciones hasta cargar el CSV oficial MINSAL del período seleccionado."
    st.markdown(f"""
    <div style="background:#FFFBEB;border:1px solid #FDE68A;border-left:5px solid #F59E0B;border-radius:0 10px 10px 0;padding:18px 20px;margin:14px 0;color:#92400E;line-height:1.6">
        <div style="font-size:15px;font-weight:800;margin-bottom:4px">⚠️ Período sin carga oficial MINSAL</div>
        <div style="font-size:13px">Aún no se han cargado datos para <strong>{year} · {pinfo['label']} · {pinfo['periodo']}</strong>.{texto_detalle}</div>
    </div>
    """, unsafe_allow_html=True)

def apply_period_context(year: int = None, reporte_id: str = None):
    """Actualiza ESTABLECIMIENTOS solo si existe carga del período. No arrastra datos de otros períodos."""
    year = year or st.session_state.get("selected_year", 2026)
    reporte_id = reporte_id or st.session_state.get("selected_report", "R1")
    ESTABLECIMIENTOS.clear()
    ESTABLECIMIENTOS.update(json.loads(json.dumps(BASE_ESTABLECIMIENTOS)))
    data = load_datos_periodos()
    period_data = data.get(str(year), {}).get(reporte_id, {}).get("establecimientos", {})
    for eid, vals in period_data.items():
        if eid in ESTABLECIMIENTOS:
            ESTABLECIMIENTOS[eid].update(vals)

def selected_year_report_controls():
    """Selector global Año + Período."""
    st.session_state.setdefault("selected_year", 2026)
    st.session_state.setdefault("selected_report", "R1")
    years = list(range(2026, 2031))
    rids = [p["id"] for p in PERIODOS]
    c1, c2, c3 = st.columns([1, 2, 4])
    with c1:
        year = st.selectbox("Año", years, index=years.index(st.session_state.selected_year), key="sel_year_widget")
    with c2:
        rid = st.selectbox("Período de trabajo", rids, index=rids.index(st.session_state.selected_report), format_func=lambda x: f"{get_periodo_info(x)['label']} — {get_periodo_info(x)['periodo']}", key="sel_report_widget")
    if year != st.session_state.selected_year or rid != st.session_state.selected_report:
        st.session_state.selected_year = year
        st.session_state.selected_report = rid
        apply_period_context(year, rid)
        st.rerun()
    apply_period_context(year, rid)
    with c3:
        estado = "✅ Con carga oficial" if periodo_tiene_datos(year, rid) else "⚠️ Sin carga oficial"
        color = "#166534" if periodo_tiene_datos(year, rid) else "#92400E"
        st.markdown(f"<div style='padding-top:28px;font-size:12px;color:#64748b'>Base activa: <b>{periodo_display(year, rid)}</b> &nbsp;·&nbsp; <b style='color:{color}'>{estado}</b></div>", unsafe_allow_html=True)

def parse_csv_minsal(file_bytes) -> dict:
    """
    Parsea el CSV oficial MINSAL con columnas:
    Codigo DEIS | RUT | Establecimiento | Servicio de Salud |
    Denominador | Numerador | % Trato Directo 2026 |
    % Trato Directo periodo equivalente 2025 | Brecha vs meta |
    Variacion vs 2025 | Nivel de riesgo
    Retorna dict con clave = eid_interno, valor = campos actualizados.
    """
    import io as _io
    content_str = file_bytes.decode("utf-8-sig")
    reader = csv.DictReader(_io.StringIO(content_str), delimiter=";")
    updates = {}
    errors = []
    rows_found = 0

    for row in reader:
        # Normalizar claves (strip espacios)
        row = {k.strip(): v.strip() for k, v in row.items() if k}
        deis = row.get("Codigo DEIS","").strip().zfill(6)
        eid = DEIS_MAP.get(deis)
        if not eid:
            # Intentar por nombre
            nombre_csv = row.get("Establecimiento","").lower()
            for k, e in ESTABLECIMIENTOS.items():
                if e["nombre"].lower() in nombre_csv or nombre_csv in e["nombre"].lower():
                    eid = k; break

        if not eid:
            errors.append(f"DEIS {deis} no reconocido: {row.get('Establecimiento','')}")
            continue

        rows_found += 1
        try:
            def parse_float(v):
                v = v.replace(",",".").replace(" ","").replace("$","").replace("%","")
                return float(v) if v else 0.0

            pct_2026 = parse_float(row.get("% Trato Directo 2026", "0"))
            pct_2025 = parse_float(row.get("% Trato Directo periodo equivalente 2025", "0"))
            brecha   = parse_float(row.get("Brecha vs meta", "0"))
            variacion= parse_float(row.get("Variacion vs 2025", "0"))
            den      = int(float(row.get("Denominador","0").replace(",",".").replace(" ","")))
            num      = int(float(row.get("Numerador","0").replace(",",".").replace(" ","")))
            nivel_raw= row.get("Nivel de riesgo","verde").strip().lower()
            nivel    = nivel_raw if nivel_raw in ["rojo","amarillo","verde"] else "verde"

            updates[eid] = {
                "pct_2026":  pct_2026,
                "pct_2025":  pct_2025,
                "brecha":    brecha,
                "variacion": variacion,
                "denominador": den,
                "numerador":   num,
                "nivel":       nivel,
                "rut":  row.get("RUT", ESTABLECIMIENTOS[eid].get("rut","")),
                "codigo_deis": deis,
                "ultima_actualizacion": datetime.datetime.now().isoformat()[:10],
                "fuente_csv": row.get("Establecimiento",""),
            }
        except Exception as ex:
            errors.append(f"{eid}: {ex}")

    return updates, errors, rows_found

# Cargar datos actualizados al iniciar
load_datos_minsal()
PERIODOS = [
    {"id":"R1","label":"1° Reporte","periodo":"Enero–Marzo 2026",     "fecha_limite":"2026-07-31","fecha_txt":"31 Jul 2026"},
    {"id":"R2","label":"2° Reporte","periodo":"Abril–Junio 2026",     "fecha_limite":"2026-08-31","fecha_txt":"31 Ago 2026"},
    {"id":"R3","label":"3° Reporte","periodo":"Jul–Sep 2026",         "fecha_limite":"2026-11-30","fecha_txt":"30 Nov 2026"},
    {"id":"R4","label":"4° Reporte","periodo":"Oct–Dic 2026",         "fecha_limite":"2027-02-28","fecha_txt":"28 Feb 2027"},
]
CAUSALES = ["Proveedor único / exclusividad","Emergencia o urgencia debidamente calificada","Licitación desierta (2° vez)","Confidencialidad / seguridad nacional","Continuidad de servicios en ejecución","Derechos de propiedad intelectual (software/licencias)","Equipamiento especializado sin equivalente","Convenio con otro organismo público (CENABAST, etc.)","Compra ágil (< 30 UTM)","Otra causal — especificar en observaciones"]

def _default_users():
    raw = {
        "admin":         ("Administrador SSMOCC",            "admin",          "abastecimiento@ssmocc.cl",         None,             "Admin2026*"),
        "bayron":        ("Bayron Retamal González",         "admin",          "bayron.retamal@ssmocc.cl",         None,             "Ssmoc2026*"),
        "traumatologico":("Miguel Jara",                    "establecimiento","miguel.jara@intraumatologico.cl",  "traumatologico", "Trauma2026*"),
        "direccion":     ("Referente Dir. SSMOCC",           "establecimiento","abast.direccion@ssmocc.cl",        "direccion",      "Dir2026*"),
        "felix_bulnes":  ("Carolina Castro",                "establecimiento","carolina.castroj@redsalud.gob.cl", "felix_bulnes",   "Felix2026*"),
        "san_juan":      ("Rodrigo Bravo Gajardo",          "establecimiento","rodrigo.bravog@redsalud.gob.cl",   "san_juan",       "Sjd2026*"),
        "crs_allende":   ("Eric Cubillo Antúnez",           "establecimiento","eric.cubillo@redsalud.gob.cl",     "crs_allende",    "Crs2026*"),
        "melipilla":     ("María de los Ángeles Morales",   "establecimiento","maria.moralesj@redsalud.gob.cl",   "melipilla",      "Meli2026*"),
        "penaflor":      ("Gissela Salvo",                  "establecimiento","gissela.salvo@redsalud.gob.cl",    "penaflor",       "Pen2026*"),
        "curacavi":      ("Pablo Yévenes Olivares",         "establecimiento","pablo.yevenes@redsalud.gob.cl",    "curacavi",       "Cura2026*"),
        "talagante":     ("María Andrea Villegas Albarrán", "establecimiento","mandrea.villegas@redsalud.gob.cl", "talagante",      "Tala2026*"),
    }
    return {uid:{"nombre":n,"rol":r,"email":e,"establecimiento":est,"password_hash":_hash(p),"activo":True} for uid,(n,r,e,est,p) in raw.items()}

def load_users():
    data = _load_json_persistent(USERS_FILE, "users.json", None)
    if data is None:
        u = _default_users()
        save_users(u)
        return u
    return data

def save_users(u):
    _save_json_persistent(USERS_FILE, "users.json", u)

def load_reports():
    return _load_json_persistent(REPORTS_FILE, "reports.json", [])

def save_reports(r):
    _save_json_persistent(REPORTS_FILE, "reports.json", r)
def upsert_report(data):
    rpts=load_reports()
    for i,r in enumerate(rpts):
        if r["establecimiento_id"]==data["establecimiento_id"] and r["reporte_id"]==data["reporte_id"]:
            rpts[i]=data; save_reports(rpts); return
    rpts.append(data); save_reports(rpts)
def delete_report(eid,rid):
    save_reports([r for r in load_reports() if not(r["establecimiento_id"]==eid and r["reporte_id"]==rid)])
def authenticate(u,p):
    users=load_users(); usr=users.get(u.strip().lower())
    if usr and usr.get("activo") and usr["password_hash"]==_hash(p): return {**usr,"username":u.strip().lower()}
    return None

def get_logo():
    p=Path(__file__).parent/"logo_ssmocc.jpg"
    if p.exists(): return base64.b64encode(p.read_bytes()).decode()
    return None
LOGO=get_logo()
def logo_img(h=38):
    if LOGO: return f'<img src="data:image/jpeg;base64,{LOGO}" style="height:{h}px;border-radius:3px">'
    return '<span style="font-size:26px">🏥</span>'

# Session
for k,v in [("user",None),("page","dashboard")]:
    if k not in st.session_state: st.session_state[k]=v

# ═══════════════════════════════════════════════════════════════════════
# LOGIN
# ═══════════════════════════════════════════════════════════════════════
def page_login():
    st.markdown(f"""
    <div style="text-align:center;margin:50px auto 24px;max-width:400px">
        {logo_img(90)}
        <div style="font-size:22px;font-weight:700;color:#1F3864;margin-top:12px">Monitor Trato Directo</div>
        <div style="font-size:13px;color:#64748b;margin-top:4px">Servicio de Salud Metropolitano Occidente</div>
        <div style="font-size:11px;color:#94a3b8;margin-top:2px">Lineamiento MINSAL v1.0 · Junio 2026</div>
    </div>""", unsafe_allow_html=True)
    _,col,_ = st.columns([1,1,1])
    with col:
        with st.form("lf"):
            u=st.text_input("👤 Usuario",placeholder="Ingrese su usuario")
            p=st.text_input("🔒 Contraseña",type="password",placeholder="Ingrese su contraseña")
            if st.form_submit_button("Ingresar",use_container_width=True,type="primary"):
                usr=authenticate(u,p)
                if usr: st.session_state.user=usr; st.session_state.page="dashboard"; st.rerun()
                else: st.error("Usuario o contraseña incorrectos.")
        st.markdown('<div style="background:#EFF6FF;border:1px solid #BFDBFE;border-radius:8px;padding:11px 14px;margin-top:12px;font-size:12px;color:#1E40AF">Acceso restringido — sistema exclusivo para referentes de abastecimiento SSMOCC.</div>', unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════════════
# TOPBAR + NAV (reemplaza al sidebar)
# ═══════════════════════════════════════════════════════════════════════
def render_topbar():
    user=st.session_state.user
    reports=load_reports()
    r1_req=len([e for e in ESTABLECIMIENTOS.values() if e["nivel"] in ["rojo","amarillo"]])
    r1_done=len([r for r in reports if r.get("reporte_id")=="R1" and r.get("estado")=="enviado"])
    pct=int(r1_done/r1_req*100) if r1_req else 0
    col_pct="#4ade80" if r1_done==r1_req else "#fbbf24" if r1_done>0 else "#f87171"
    estab=ESTABLECIMIENTOS.get(user.get("establecimiento"),{})
    rol_txt="🔑 Admin" if user["rol"]=="admin" else f"🏥 {estab.get('nombre_corto','')}"

    st.markdown(f"""
    <div class="topbar">
      <div class="topbar-inner">
        <div class="topbar-logo">{logo_img(38)}</div>
        <div class="topbar-title">
          <div class="t1">Monitor Trato Directo — SSMOCC 2026</div>
          <div class="t2">Lineamiento MINSAL v1.0 · Subsecretaría de Redes Asistenciales</div>
        </div>
        <div style="padding:6px 14px;border-left:1px solid rgba(255,255,255,.15);text-align:center;min-width:120px">
          <div style="font-size:10px;color:rgba(255,255,255,.4);text-transform:uppercase;letter-spacing:.06em">1° Reporte</div>
          <div style="font-size:16px;font-weight:700;color:{col_pct}">{r1_done}/{r1_req}</div>
          <div style="background:rgba(255,255,255,.15);border-radius:3px;height:3px;margin-top:4px">
            <div style="width:{pct}%;height:100%;background:{col_pct};border-radius:3px"></div>
          </div>
        </div>
        <div class="topbar-user">
          <div class="u1">{user['nombre']}</div>
          <div class="u2">{rol_txt}</div>
        </div>
      </div>
    </div>""", unsafe_allow_html=True)

    # Nav pills
    nav=[("📊","Dashboard","dashboard"),("📋","Mis reportes","mis_reportes")]
    if user["rol"]=="admin":
        nav+=[("📁","Todos los reportes","todos_reportes"),("📤","Exportar MINSAL","exportar"),("📰","Boletines","boletines"),("📈","Histórico","historico"),("📂","Carga MINSAL","actualizar_datos"),("👥","Usuarios","usuarios"),("⚙️","Config.","configuracion")]
    nav.append(("🚪","Cerrar sesión","__logout__"))

    cols=st.columns(len(nav))
    for i,(icon,label,pid) in enumerate(nav):
        active=st.session_state.page==pid
        with cols[i]:
            if st.button(f"{icon} {label}",key=f"nav_{pid}",use_container_width=True,
                        type="primary" if active else "secondary"):
                if pid=="__logout__":
                    st.session_state.user=None; st.session_state.page="dashboard"; st.rerun()
                else:
                    st.session_state.page=pid; st.rerun()
    st.markdown("<hr style='margin:0 0 12px;border-color:#e2e8f0'>",unsafe_allow_html=True)
    selected_year_report_controls()
    st.markdown('''<div style="text-align:right;font-size:10px;color:#94a3b8;margin-top:-10px;padding-right:4px;margin-bottom:4px">
        Desarrollado por <strong style="color:#64748b">Bayron Retamal González</strong> &nbsp;·&nbsp; Subdirección RRFF &nbsp;·&nbsp; SSMOCC 2026
    </div>''', unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════════════
# HELPERS UI
# ═══════════════════════════════════════════════════════════════════════
def page_header(title,sub=""):
    st.markdown(f"""
    <div style="background:#1F3864;padding:12px 20px;border-radius:10px;margin-bottom:16px;display:flex;align-items:center;gap:12px">
        {logo_img(36)}
        <div style="width:4px;background:#C0392B;border-radius:2px;align-self:stretch;min-height:30px"></div>
        <div>
            <div style="font-size:16px;font-weight:700;color:white">{title}</div>
            <div style="font-size:11px;color:rgba(255,255,255,.5);margin-top:2px">{sub}</div>
        </div>
    </div>""", unsafe_allow_html=True)

def kpi_card(col,label,value,note,color="#1F3864"):
    col.markdown(f"""
    <div class="kcard" style="border-top-color:{color}">
        <div class="lbl">{label}</div>
        <div class="val" style="color:{color}">{value}</div>
        <div class="note">{note}</div>
    </div>""", unsafe_allow_html=True)

def nivel_pill(nivel):
    cls={"rojo":"nr","amarillo":"na","verde":"nv"}.get(nivel,"nv")
    return f'<span class="{cls}">{nivel.capitalize()}</span>'

def cal_cards(reports, eid_filter=None):
    hoy=datetime.date.today()
    req=len([e for e in ESTABLECIMIENTOS.values() if e["nivel"] in ["rojo","amarillo"]])
    cols=st.columns(4)
    for i,p in enumerate(PERIODOS):
        fl=datetime.date.fromisoformat(p["fecha_limite"])
        dias=(fl-hoy).days
        if eid_filter:
            r=next((x for x in reports if x.get("establecimiento_id")==eid_filter and x.get("reporte_id")==p["id"]),None)
            est=r.get("estado","pendiente") if r else "pendiente"
            if est=="enviado":     bg,tc,bc,tag="#F0FDF4","#166534","#BBF7D0","✅ ENVIADO"
            elif est=="borrador":  bg,tc,bc,tag="#FFFBEB","#92400E","#FDE68A","📝 BORRADOR"
            elif dias<0:           bg,tc,bc,tag="#FEF2F2","#991B1B","#FECACA","⛔ VENCIDO"
            elif dias<=14:         bg,tc,bc,tag="#FEF2F2","#991B1B","#FECACA",f"⚠️ {dias}d"
            elif dias<=30:         bg,tc,bc,tag="#FFFBEB","#92400E","#FDE68A",f"🔔 {dias}d"
            else:                  bg,tc,bc,tag="#F8FAFC","#475569","#E2E8F0",f"📅 {dias}d"
            sub=f"Estado: <b>{est.upper()}</b>"
        else:
            env=len([r for r in reports if r.get("reporte_id")==p["id"] and r.get("estado")=="enviado"])
            if env==req:           bg,tc,bc,tag="#F0FDF4","#166534","#BBF7D0","✅ COMPLETO"
            elif dias<0:           bg,tc,bc,tag="#FEF2F2","#991B1B","#FECACA","⛔ VENCIDO"
            elif dias<=14:         bg,tc,bc,tag="#FEF2F2","#991B1B","#FECACA",f"⚠️ {dias}d"
            elif dias<=30:         bg,tc,bc,tag="#FFFBEB","#92400E","#FDE68A",f"🔔 {dias}d"
            else:                  bg,tc,bc,tag="#F8FAFC","#475569","#E2E8F0",f"📅 {dias}d"
            sub=f"{env}/{req} enviados"
        with cols[i]:
            st.markdown(f"""
            <div style="background:{bg};border:1px solid {bc};border-top:3px solid {tc};
                        border-radius:8px;padding:10px;text-align:center;height:110px;display:flex;flex-direction:column;justify-content:center">
                <div style="font-size:14px;font-weight:800;color:{tc}">{p['label']}</div>
                <div style="font-size:10px;color:{tc};margin:2px 0">{p['periodo']}</div>
                <div style="font-size:11px;font-weight:600;color:{tc}">{p['fecha_txt']}</div>
                <div style="background:rgba(0,0,0,.07);border-radius:4px;padding:2px 6px;margin-top:5px;font-size:10px;color:{tc};font-weight:700">{tag}</div>
                <div style="font-size:10px;color:{tc};margin-top:2px">{sub}</div>
            </div>""", unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════════════
# PÁGINA DASHBOARD
# ═══════════════════════════════════════════════════════════════════════
def pg_dashboard():
    import plotly.graph_objects as go
    import pandas as pd
    year = st.session_state.get("selected_year", 2026)
    rid = st.session_state.get("selected_report", "R1")
    apply_period_context(year, rid)
    if not periodo_tiene_datos(year, rid):
        page_header("Monitoreo Trato Directo — Red SSMOCC", "Dashboard ejecutivo por año y período")
        mensaje_periodo_sin_datos(year, rid)
        st.info("Ingrese al módulo **Carga MINSAL** para cargar el CSV oficial del período. Una vez cargado, se habilitarán automáticamente KPIs, semáforos, ranking, boletín y exportaciones.")
        return
    user=st.session_state.user
    es_admin=user["rol"]=="admin"
    eid_user=user.get("establecimiento")
    reports=load_reports()

    # ── VISTA ESTABLECIMIENTO ─────────────────────────────────────────
    if not es_admin and eid_user:
        e=ESTABLECIMIENTOS.get(eid_user,{})
        nivel=e.get("nivel","verde")
        nc={"rojo":"#E24B4A","amarillo":"#F59E0B","verde":"#22C55E"}.get(nivel,"#22C55E")
        nt={"rojo":"#991B1B","amarillo":"#92400E","verde":"#166534"}.get(nivel,"#166534")
        nb={"rojo":"#FEF2F2","amarillo":"#FFFBEB","verde":"#F0FDF4"}.get(nivel,"#F0FDF4")
        nbc={"rojo":"#FECACA","amarillo":"#FDE68A","verde":"#BBF7D0"}.get(nivel,"#BBF7D0")
        va=f"+{e.get('variacion',0):.2f}" if e.get("variacion",0)>0 else f"{e.get('variacion',0):.2f}"

        page_header(f"Mi establecimiento — {e.get('nombre_corto','')}",
                    "Datos oficiales MINSAL · Período enero–junio 2026 · Meta institucional ≤ 16%")

        c1,c2,c3,c4=st.columns(4)
        kpi_card(c1,"Numerador CLP",f"${e.get('numerador',0)/1e9:.1f} MM","Monto TD recepción conforme","#0C447C")
        kpi_card(c2,"Denominador CLP",f"${e.get('denominador',0)/1e9:.1f} MM","Total todas las modalidades","#0C447C")
        kpi_card(c3,"% TD 2026",f"{e.get('pct_2026',0):.2f}%",f"Brecha {'+' if e.get('brecha',0)>0 else ''}{e.get('brecha',0):.2f} pp vs meta",nt)
        kpi_card(c4,"% TD 2025 mismo período",f"{e.get('pct_2025',0):.2f}%",f"Variación: {va} pp","#0C447C")

        acciones={"rojo":"Elaborar plan de acción · Reunión técnica de seguimiento · Monitoreo mensual · Ingresar reporte con causas y compromisos.","amarillo":"Analizar causas del resultado · Identificar compras para migrar a mecanismos competitivos · Ingresar reporte con medidas implementadas.","verde":"Mantener buenas prácticas · No se requiere remitir antecedentes a la Subsecretaría."}
        titulo_e={"rojo":"Requiere plan de acción urgente","amarillo":"Monitoreo reforzado","verde":"Mantener buenas prácticas"}
        st.markdown(f"""
        <div style="background:{nb};border-left:5px solid {nc};border:1px solid {nbc};padding:12px 16px;border-radius:8px;margin:12px 0">
            <div style="font-size:13px;font-weight:700;color:{nt}">Nivel {nivel.upper()} — {titulo_e[nivel]}</div>
            <div style="font-size:12px;color:{nt};margin-top:4px;line-height:1.6">{acciones[nivel]}</div>
        </div>""", unsafe_allow_html=True)

        col_g,col_r=st.columns([1.2,1])
        with col_g:
            st.markdown('<div style="font-size:12px;font-weight:600;color:#1F3864;margin-bottom:6px">Comparación % TD — 2025 vs 2026</div>',unsafe_allow_html=True)
            fig=go.Figure()
            fig.add_trace(go.Bar(name="% TD 2025",x=["Mismo período 2025"],y=[e.get("pct_2025",0)],marker_color="#CBD5E1",text=[f"{e.get('pct_2025',0):.1f}%"],textposition="outside"))
            fig.add_trace(go.Bar(name="% TD 2026",x=["Acumulado 2026"],y=[e.get("pct_2026",0)],marker_color=nc,text=[f"{e.get('pct_2026',0):.1f}%"],textposition="outside"))
            fig.add_hline(y=16,line_dash="dash",line_color="#1F3864",line_width=1.5,annotation_text="Meta 16%",annotation_position="top right")
            fig.update_layout(height=200,margin=dict(l=0,r=60,t=20,b=20),plot_bgcolor="white",paper_bgcolor="white",showlegend=True,legend=dict(orientation="h",y=1.1,x=0),yaxis=dict(range=[0,max(e.get("pct_2026",0)+15,30)],gridcolor="#f1f5f9"),font=dict(size=11,family="Arial"))
            st.plotly_chart(fig,use_container_width=True)
        with col_r:
            st.markdown('<div style="font-size:12px;font-weight:600;color:#1F3864;margin-bottom:6px">Datos MINSAL oficiales</div>',unsafe_allow_html=True)
            st.markdown(f"""
            <div style="background:white;border:0.5px solid #e2e8f0;border-radius:8px;padding:12px">
                <div style="display:grid;grid-template-columns:1fr 1fr;gap:8px;font-size:12px">
                    <div><span style="color:#64748b">Código DEIS</span><br><b>{e.get('codigo_deis','')}</b></div>
                    <div><span style="color:#64748b">RUT</span><br><b>{e.get('rut','')}</b></div>
                    <div><span style="color:#64748b">Numerador</span><br><b>${e.get('numerador',0):,.0f}</b></div>
                    <div><span style="color:#64748b">Denominador</span><br><b>${e.get('denominador',0):,.0f}</b></div>
                    <div><span style="color:#64748b">% TD 2026</span><br><b style="color:{nt}">{e.get('pct_2026',0):.2f}%</b></div>
                    <div><span style="color:#64748b">Brecha vs meta</span><br><b style="color:{nt}">{'+' if e.get('brecha',0)>0 else ''}{e.get('brecha',0):.2f} pp</b></div>
                </div>
            </div>""", unsafe_allow_html=True)

        st.markdown('<div style="font-size:12px;font-weight:600;color:#1F3864;margin:14px 0 8px">Estado de mis reportes</div>',unsafe_allow_html=True)
        cal_cards(reports, eid_filter=eid_user)

        st.markdown('<div style="font-size:12px;font-weight:600;color:#1F3864;margin:14px 0 8px">📅 Plazos — Lineamiento MINSAL v1.0</div>',unsafe_allow_html=True)
        st.markdown('<div style="background:#EFF6FF;border:1px solid #BFDBFE;border-radius:6px;padding:9px 14px;font-size:11px;color:#1E40AF">Los antecedentes se remiten de manera consolidada por el SSMOCC, incluyendo únicamente establecimientos en categoría Amarilla y Roja.</div>',unsafe_allow_html=True)

        if nivel in ["rojo","amarillo"]:
            st.markdown("<br>",unsafe_allow_html=True)
            if st.button("📋  Ingresar mi reporte →",type="primary"):
                st.session_state.page="mis_reportes"; st.rerun()
        return

    # ── VISTA ADMIN ───────────────────────────────────────────────────
    page_header("Monitoreo Trato Directo — Red SSMOCC 2026",
                "Lineamiento MINSAL v1.0 · Junio 2026 · Fuente: ChileCompra — OC aceptadas con recepción conforme, monto neto CLP")

    total_num=sum(e["numerador"] for e in ESTABLECIMIENTOS.values())
    total_den=sum(e["denominador"] for e in ESTABLECIMIENTOS.values())
    pct_s=round(total_num/total_den*100,2)
    rojos=[e for e in ESTABLECIMIENTOS.values() if e["nivel"]=="rojo"]
    amarillos=[e for e in ESTABLECIMIENTOS.values() if e["nivel"]=="amarillo"]
    verdes=[e for e in ESTABLECIMIENTOS.values() if e["nivel"]=="verde"]
    r1_req=len(rojos)+len(amarillos)
    r1_done=len([r for r in reports if r.get("reporte_id")=="R1" and r.get("estado")=="enviado"])

    c1,c2,c3,c4=st.columns(4)
    kpi_card(c1,"Numerador SSMOCC",f"${total_num/1e12:.2f} MMM","TD con recepción conforme","#0C447C")
    kpi_card(c2,"Denominador SSMOCC",f"${total_den/1e12:.2f} MMM","Todas las modalidades","#0C447C")
    kpi_card(c3,"% TD SSMOCC 2026",f"{pct_s:.2f}%",f"Meta ≤ 16% · Brecha +{pct_s-16:.2f} pp","#A32D2D")
    kpi_card(c4,"1° Reporte (31 Jul)",f"{r1_done}/{r1_req}","Establecimientos enviados","#0F6E56")

    st.markdown("<br>",unsafe_allow_html=True)
    col_g,col_s=st.columns([3,1.2])
    with col_g:
        st.markdown('<div style="font-size:12px;font-weight:600;color:#1F3864;margin-bottom:6px">% TD por establecimiento — comparación 2025 vs 2026</div>',unsafe_allow_html=True)
        df=pd.DataFrame([{"Establecimiento":e["nombre_corto"],"2025":e["pct_2025"],"2026":e["pct_2026"],"nivel":e["nivel"]} for e in ESTABLECIMIENTOS.values()]).sort_values("2026",ascending=True)
        cmap={"rojo":"#E24B4A","amarillo":"#F59E0B","verde":"#22C55E"}
        fig=go.Figure()
        fig.add_trace(go.Bar(name="% TD 2025",y=df["Establecimiento"],x=df["2025"],orientation="h",marker_color="#CBD5E1",opacity=0.7))
        fig.add_trace(go.Bar(name="% TD 2026",y=df["Establecimiento"],x=df["2026"],orientation="h",marker_color=[cmap[n] for n in df["nivel"]]))
        fig.add_vline(x=16,line_dash="dash",line_color="#1F3864",line_width=1.5,annotation_text="Meta 16%",annotation_font_color="#1F3864")
        fig.update_layout(barmode="overlay",height=280,margin=dict(l=0,r=20,t=10,b=20),legend=dict(orientation="h",yanchor="bottom",y=1.02,x=0,font=dict(size=10)),plot_bgcolor="white",paper_bgcolor="white",xaxis=dict(range=[0,70],gridcolor="#f1f5f9"),font=dict(size=10,family="Arial"))
        st.plotly_chart(fig,use_container_width=True)
    with col_s:
        st.markdown('<div style="font-size:12px;font-weight:600;color:#1F3864;margin-bottom:6px">Semáforo de riesgo</div>',unsafe_allow_html=True)
        for nv,estabs,bg,tc,bc in [("🔴 Rojo",rojos,"#FEF2F2","#991B1B","#FECACA"),("🟡 Amarillo",amarillos,"#FFFBEB","#92400E","#FDE68A"),("🟢 Verde",verdes,"#F0FDF4","#166534","#BBF7D0")]:
            names="<br>".join(f"• {e['nombre_corto']}" for e in estabs)
            st.markdown(f'<div style="background:{bg};border:1px solid {bc};border-radius:8px;padding:8px 12px;margin-bottom:6px"><div style="font-size:11px;font-weight:700;color:{tc}">{nv} ({len(estabs)})</div><div style="font-size:10px;color:{tc};margin-top:3px;line-height:1.6">{names}</div></div>',unsafe_allow_html=True)

    st.markdown("<hr style='border-color:#f1f5f9;margin:12px 0'>",unsafe_allow_html=True)

    # Casos críticos
    st.markdown('<div style="font-size:12px;font-weight:600;color:#1F3864;margin-bottom:8px">Establecimientos en nivel rojo</div>',unsafe_allow_html=True)
    cols_r=st.columns(len(rojos))
    for i,e in enumerate(sorted(rojos,key=lambda x:-x["pct_2026"])):
        va=f"+{e['variacion']:.1f}" if e["variacion"]>0 else f"{e['variacion']:.1f}"
        with cols_r[i]:
            st.markdown(f'<div style="background:white;border:1px solid #FECACA;border-top:4px solid #E24B4A;border-radius:8px;padding:12px"><div style="font-size:11px;font-weight:600;color:#991B1B">{e["nombre_corto"]}</div><div style="font-size:26px;font-weight:800;color:#A32D2D;line-height:1.1">{e["pct_2026"]:.1f}%</div><div style="font-size:10px;color:#DC2626;margin-top:3px">Brecha: +{e["brecha"]:.1f} pp · Var.: {va} pp</div></div>',unsafe_allow_html=True)

    st.markdown("<hr style='border-color:#f1f5f9;margin:12px 0'>",unsafe_allow_html=True)

    # Tabla estado
    st.markdown('<div style="font-size:12px;font-weight:600;color:#1F3864;margin-bottom:6px">Estado de reportes</div>',unsafe_allow_html=True)
    rows=[]
    for eid,e in ESTABLECIMIENTOS.items():
        if e["nivel"]=="verde": continue
        row={"Establecimiento":e["nombre_corto"],"Nivel":f"{'🔴' if e['nivel']=='rojo' else '🟡'} {e['nivel'].capitalize()}","% TD":f"{e['pct_2026']:.1f}%"}
        for p in PERIODOS:
            r=next((x for x in reports if x.get("establecimiento_id")==eid and x.get("reporte_id")==p["id"]),None)
            row[p["label"]]="✅ Enviado" if r and r.get("estado")=="enviado" else "📝 Borrador" if r else "⬜ Pendiente"
        rows.append(row)
    st.dataframe(pd.DataFrame(rows),use_container_width=True,hide_index=True)

    st.markdown("<hr style='border-color:#f1f5f9;margin:12px 0'>",unsafe_allow_html=True)
    st.markdown('<div style="font-size:12px;font-weight:600;color:#1F3864;margin-bottom:8px">📅 Calendario de reportes — Lineamiento MINSAL v1.0</div>',unsafe_allow_html=True)
    cal_cards(reports)
    st.markdown('<div style="background:#EFF6FF;border:1px solid #BFDBFE;border-radius:6px;padding:9px 14px;font-size:11px;color:#1E40AF;margin-top:8px"><strong>Nota MINSAL:</strong> Los antecedentes se remiten consolidados por el Servicio de Salud, incluyendo únicamente establecimientos en categoría Amarilla y Roja. Las fechas pueden ajustarse según disponibilidad de ChileCompra.</div>',unsafe_allow_html=True)

    with st.expander("📊 Contexto nacional — 224 establecimientos MINSAL"):
        nc1,nc2,nc3,nc4=st.columns(4)
        nc1.metric("Numerador nacional","$238,1 MMM"); nc2.metric("Denominador nacional","$1.434,9 MMM")
        nc3.metric("% TD nacional 2026","16,6%"); nc4.metric("Variación vs 2025","−4,3 pp")
        st.markdown("El **Instituto Traumatológico SSMOCC** (62,2%) ocupa el **5° lugar nacional** entre establecimientos con mayor % TD.")


# ═══════════════════════════════════════════════════════════════════════
# PÁGINA MIS REPORTES
# ═══════════════════════════════════════════════════════════════════════

def pg_todos_reportes():
    if st.session_state.user["rol"]!="admin": st.error("Solo administradores."); return
    import pandas as pd
    page_header("Todos los reportes","Vista consolidada y gestión — solo administradores")
    year = st.session_state.get("selected_year", 2026)
    rid = st.session_state.get("selected_report", "R1")
    apply_period_context(year, rid)
    if not periodo_tiene_datos(year, rid):
        mensaje_periodo_sin_datos(year, rid)
        return
    reports=load_reports()
    c1,c2,c3=st.columns(3)
    fp=c1.selectbox("Período",["Todos"]+[p["id"] for p in PERIODOS],format_func=lambda x:"Todos" if x=="Todos" else next(p["label"]+" — "+p["periodo"] for p in PERIODOS if p["id"]==x))
    fn=c2.selectbox("Nivel",["Todos","Rojo","Amarillo","Verde"])
    fe=c3.selectbox("Estado",["Todos","Enviado","Borrador","Pendiente"])
    st.subheader("Estado consolidado")
    rows=[]
    for eid,e in ESTABLECIMIENTOS.items():
        if e["nivel"]=="verde": continue
        row={"Establecimiento":e["nombre_corto"],"Nivel":e["nivel"].capitalize(),"% TD":f"{e['pct_2026']:.1f}%"}
        for p in PERIODOS:
            r=next((x for x in reports if x.get("establecimiento_id")==eid and x.get("reporte_id")==p["id"]),None)
            row[p["label"]]="✅ Enviado" if r and r.get("estado")=="enviado" else "📝 Borrador" if r else "⬜ Pendiente"
        rows.append(row)
    st.dataframe(pd.DataFrame(rows),use_container_width=True,hide_index=True)
    st.divider()
    st.subheader("Detalle de reportes")
    filtered=reports
    if fp!="Todos": filtered=[r for r in filtered if r.get("reporte_id")==fp]
    if fn!="Todos": filtered=[r for r in filtered if r.get("nivel_riesgo","").lower()==fn.lower()]
    if fe!="Todos": filtered=[r for r in filtered if r.get("estado","").lower()==fe.lower()]
    if not filtered: st.info("No hay reportes con los filtros seleccionados."); return
    for r in sorted(filtered,key=lambda x:(x.get("reporte_id",""),x.get("nivel_riesgo",""))):
        estado=r.get("estado","borrador")
        icon="✅" if estado=="enviado" else "📝"
        with st.expander(f"{icon} {r.get('establecimiento_nombre','')[:45]} — {r.get('periodo_label','')} — {estado.upper()}"):
            ca,cb,cc,cd=st.columns(4)
            ca.metric("Nivel",r.get("nivel_riesgo","").capitalize()); cb.metric("% TD 2026",f"{r.get('pct_2026',0):.2f}%")
            cc.metric("% TD período",f"{r.get('pct_per',0):.2f}%"); cd.metric("N° procesos TD",r.get("n_proc","—"))
            if r.get("causas_sel"): st.markdown("**Causales:** "+", ".join(r["causas_sel"]))
            if r.get("causas_desc"): st.markdown(f"**Descripción:** {r['causas_desc']}")
            if r.get("compromisos"): st.markdown(f"**Compromisos:** {r['compromisos']}")
            st.markdown(f"**Responsable:** {r.get('resp_nombre','')} · {r.get('resp_cargo','')} · {r.get('resp_email','')}  \n**Meta próximo período:** {r.get('meta_prox',16):.1f}% · **Fecha:** {r.get('fecha_comp','')}  \n**Ingresado:** {r.get('usuario','')} · {r.get('fecha_ingreso','')[:16]}")
            col1,col2,_=st.columns([1,1,3])
            with col1:
                if estado=="borrador":
                    if st.button("✅ Marcar enviado",key=f"s_{r['establecimiento_id']}_{r['reporte_id']}"):
                        r["estado"]="enviado"; upsert_report(r); st.rerun()
                else:
                    if st.button("↩️ A borrador",key=f"b_{r['establecimiento_id']}_{r['reporte_id']}"):
                        r["estado"]="borrador"; upsert_report(r); st.rerun()
            with col2:
                if st.button("🗑️ Eliminar",key=f"d_{r['establecimiento_id']}_{r['reporte_id']}",type="secondary"):
                    delete_report(r["establecimiento_id"],r["reporte_id"])
                    st.warning(f"Reporte eliminado."); st.rerun()


# ═══════════════════════════════════════════════════════════════════════
# PÁGINA EXPORTAR
# ═══════════════════════════════════════════════════════════════════════
def pg_exportar():
    if st.session_state.user["rol"]!="admin": st.error("Solo administradores."); return
    import pandas as pd
    page_header("Exportar reporte consolidado","Formato Anexo N°1 — Lineamiento MINSAL v1.0")
    reports=load_reports()
    ps=st.selectbox("Período",options=[p["id"] for p in PERIODOS],format_func=lambda x:next(f"{p['label']} — {p['periodo']} (plazo: {p['fecha_txt']})" for p in PERIODOS if p["id"]==x))
    pinfo=next(p for p in PERIODOS if p["id"]==ps)
    inc=st.radio("Incluir",["Solo enviados","Enviados y borradores"],horizontal=True)
    r_p=[r for r in reports if r.get("reporte_id")==ps]
    if inc=="Solo enviados": r_p=[r for r in r_p if r.get("estado")=="enviado"]
    if not r_p:
        st.warning(f"No hay reportes para {pinfo['label']} con los filtros seleccionados.")
        pend=[eid for eid,e in ESTABLECIMIENTOS.items() if e["nivel"] in ["rojo","amarillo"]]
        for eid in pend:
            e=ESTABLECIMIENTOS[eid]; r=next((x for x in reports if x.get("establecimiento_id")==eid and x.get("reporte_id")==ps),None)
            ic="✅" if r and r.get("estado")=="enviado" else "📝" if r else "⬜"
            st.markdown(f"{ic} {e['nombre']} — {e['nivel'].capitalize()}")
        return
    rows=[]
    for r in r_p:
        eid=r["establecimiento_id"]; estab_d=ESTABLECIMIENTOS.get(eid,{})
        med=r.get("medidas",{})
        med_labels={"pac":"Actualizacion Plan Anual Compras","lic":"Inicio procesos licitatorios","cm":"Migracion Convenio Marco","cenabast":"Gestion CENABAST","cap":"Capacitacion equipo Ley 21.634","venc":"Control vencimiento contratos"}
        med_txt="; ".join(lbl for k,lbl in med_labels.items() if med.get(k))
        if r.get("med_desc"): med_txt=(med_txt+" | " if med_txt else "")+r.get("med_desc","")
        causas_txt="; ".join(r.get("causas_sel",[]))
        if r.get("causas_desc"): causas_txt=(causas_txt+" | " if causas_txt else "")+r.get("causas_desc","")
        rows.append({
            "Servicio de salud":"Metropolitano Occidente",
            "Establecimiento":r.get("establecimiento_nombre",""),
            "Nivel de Riesgo":r.get("nivel_riesgo","").capitalize(),
            "Periodo informado":r.get("periodo",""),
            "Principales causas":causas_txt,
            "Medidas implementadas":med_txt,
            "Compromisos":r.get("compromisos",""),
            "Responsable":r.get("resp_nombre","")+" - "+r.get("resp_cargo",""),
            "Fecha comprometida":r.get("fecha_comp",""),
            "Codigo DEIS":estab_d.get("codigo_deis",""),
            "RUT":estab_d.get("rut",""),
            "Pct TD 2026":r.get("pct_2026",""),
            "Pct TD 2025":r.get("pct_2025",""),
            "Brecha pp":estab_d.get("brecha",""),
            "Variacion pp":estab_d.get("variacion",""),
            "Denominador CLP":estab_d.get("denominador",""),
            "Numerador CLP":estab_d.get("numerador",""),
            "Monto TD periodo CLP":r.get("monto_td",0),
            "N procesos TD":r.get("n_proc",0),
            "Correo responsable":r.get("resp_email",""),
            "Meta proximo periodo":r.get("meta_prox",""),
            "Observaciones":r.get("obs",""),
            "Estado reporte":r.get("estado","").upper(),
            "Fecha ingreso":r.get("fecha_ingreso","")[:16],
        })
    df=pd.DataFrame(rows)
    cols_minsal=["Servicio de salud","Establecimiento","Nivel de Riesgo","Periodo informado","Principales causas","Medidas implementadas","Compromisos","Responsable","Fecha comprometida"]
    st.markdown("**Vista previa Anexo N°1 MINSAL** (columnas exactas del lineamiento):")
    st.dataframe(df[cols_minsal],use_container_width=True,hide_index=True)
    st.markdown(f"**{len(rows)} establecimiento(s)** · Generado: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}")
    buf=io.BytesIO()
    with pd.ExcelWriter(buf,engine="openpyxl") as w:
        df[cols_minsal].to_excel(w,sheet_name="Anexo N1 MINSAL",index=False)
        df.to_excel(w,sheet_name="Datos completos",index=False)
        resumen=pd.DataFrame([{"Establecimiento":e["nombre"],"Nivel":e["nivel"].capitalize(),"Pct TD 2026":e["pct_2026"],"Pct TD 2025":e["pct_2025"],"Brecha pp":e["brecha"],"Variacion pp":e["variacion"],"Denominador":e["denominador"],"Numerador":e["numerador"],"Enviado":"Si" if any(r.get("establecimiento_id")==eid and r.get("estado")=="enviado" and r.get("reporte_id")==ps for r in reports) else "No"} for eid,e in ESTABLECIMIENTOS.items()])
        resumen.to_excel(w,sheet_name="Resumen SSMOCC",index=False)
    buf.seek(0)
    fecha=datetime.datetime.now().strftime("%Y%m%d_%H%M")
    c1,c2=st.columns(2)
    c1.download_button("⬇️ Descargar Excel (.xlsx)",buf,f"SSMOCC_AnexoN1_{ps}_{fecha}.xlsx","application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",use_container_width=True,type="primary")
    c2.download_button("⬇️ Descargar CSV",df.to_csv(index=False,encoding="utf-8-sig").encode("utf-8-sig"),f"SSMOCC_AnexoN1_{ps}_{fecha}.csv","text/csv",use_container_width=True)
    pend=[eid for eid,e in ESTABLECIMIENTOS.items() if e["nivel"] in ["rojo","amarillo"] and not any(r.get("establecimiento_id")==eid and r.get("estado")=="enviado" and r.get("reporte_id")==ps for r in reports)]
    if pend:
        st.warning(f"⚠️ {len(pend)} establecimiento(s) sin enviar para {pinfo['label']}:")
        for eid in pend: e=ESTABLECIMIENTOS[eid]; st.markdown(f"- **{e['nombre']}** — {e['nivel'].capitalize()} ({e['pct_2026']:.1f}%)")


# ═══════════════════════════════════════════════════════════════════════
# PÁGINA USUARIOS
# ═══════════════════════════════════════════════════════════════════════
def pg_usuarios():
    if st.session_state.user["rol"]!="admin": st.error("Solo administradores."); return
    import pandas as pd
    page_header("Gestión de usuarios","Alta, baja y restablecimiento de contraseñas")
    users=load_users()
    tab1,tab2=st.tabs(["📋 Usuarios","➕ Crear/Editar"])
    with tab1:
        rows=[{"Usuario":uid,"Nombre":u["nombre"],"Rol":u["rol"].capitalize(),"Establecimiento":ESTABLECIMIENTOS.get(u.get("establecimiento"),{}).get("nombre_corto","— Admin —"),"Email":u.get("email",""),"Activo":"✅" if u.get("activo") else "❌"} for uid,u in users.items()]
        st.dataframe(pd.DataFrame(rows),use_container_width=True,hide_index=True)
        ca,cb=st.columns(2)
        uid_sel=ca.selectbox("Usuario",list(users.keys()),format_func=lambda x:f"{x} — {users[x]['nombre']}")
        act=users.get(uid_sel,{}).get("activo",True)
        if cb.button(f"{'🔴 Desactivar' if act else '🟢 Activar'}",use_container_width=True):
            users[uid_sel]["activo"]=not act; save_users(users); st.success("Actualizado."); st.rerun()
        with st.expander("🔑 Restablecer contraseña"):
            ur=st.selectbox("Usuario",list(users.keys()),format_func=lambda x:f"{x} — {users[x]['nombre']}",key="ur")
            np=st.text_input("Nueva contraseña",type="password",key="np")
            if st.button("Restablecer"):
                if len(np)<8: st.error("Mínimo 8 caracteres.")
                else: users[ur]["password_hash"]=_hash(np); save_users(users); st.success(f"Contraseña restablecida para '{ur}'.")
        with st.expander("📋 Credenciales por defecto"):
            st.markdown("""| Usuario | Contraseña | Referente |\n|---------|-----------|----------|\n| `admin` | `Admin2026*` | Administrador |\n| `bayron` | `Ssmoc2026*` | Bayron Retamal |\n| `san_juan` | `Sjd2026*` | Rodrigo Bravo |\n| `traumatologico` | `Trauma2026*` | Miguel Jara |\n| `felix_bulnes` | `Felix2026*` | Carolina Castro |\n| `talagante` | `Tala2026*` | M.A. Villegas |\n| `penaflor` | `Pen2026*` | Gissela Salvo |\n| `melipilla` | `Meli2026*` | M.A. Morales |\n| `curacavi` | `Cura2026*` | Pablo Yévenes |\n| `crs_allende` | `Crs2026*` | Eric Cubillo |""")
    with tab2:
        modo=st.radio("Modo",["Crear nuevo","Editar existente"],horizontal=True)
        uid_e=None; u_e={}
        if modo=="Editar existente":
            uid_e=st.selectbox("Usuario a editar",list(users.keys()),format_func=lambda x:f"{x} — {users[x]['nombre']}",key="ue"); u_e=users.get(uid_e,{})
        with st.form("fu"):
            c1,c2=st.columns(2)
            nuevo_uid=c1.text_input("Nombre de usuario") if modo=="Crear nuevo" else st.text_input("Usuario",value=uid_e,disabled=True)
            nombre=c1.text_input("Nombre completo",value=u_e.get("nombre",""))
            email=c1.text_input("Correo",value=u_e.get("email",""))
            rol=c2.selectbox("Rol",["establecimiento","admin"],index=0 if u_e.get("rol","establecimiento")=="establecimiento" else 1)
            estab_ops={"":"— Solo admin —"}; estab_ops.update({eid:e["nombre_corto"] for eid,e in ESTABLECIMIENTOS.items()})
            estab_sel=c2.selectbox("Establecimiento",list(estab_ops.keys()),format_func=lambda x:estab_ops[x],index=list(estab_ops.keys()).index(u_e.get("establecimiento","") or ""))
            pwd=c2.text_input("Contraseña"+" (vacío=no cambiar)" if modo=="Editar existente" else "",type="password")
            activo=c2.checkbox("Activo",value=u_e.get("activo",True))
            if st.form_submit_button("💾 Guardar",use_container_width=True,type="primary"):
                uid_f=(nuevo_uid if modo=="Crear nuevo" else uid_e).strip().lower()
                errs=[]
                if not uid_f: errs.append("Usuario vacío.")
                if not nombre.strip(): errs.append("Nombre requerido.")
                if modo=="Crear nuevo" and uid_f in users: errs.append(f"'{uid_f}' ya existe.")
                if modo=="Crear nuevo" and not pwd: errs.append("Contraseña requerida.")
                if errs:
                    for e in errs: st.error(e)
                else:
                    if uid_f not in users: users[uid_f]={}
                    users[uid_f].update({"nombre":nombre.strip(),"rol":rol,"email":email.strip(),"establecimiento":estab_sel or None,"activo":activo})
                    if pwd: users[uid_f]["password_hash"]=_hash(pwd)
                    save_users(users); st.success(f"✅ Usuario '{uid_f}' guardado."); st.rerun()


# ═══════════════════════════════════════════════════════════════════════
# PÁGINA CONFIGURACIÓN
# ═══════════════════════════════════════════════════════════════════════
def pg_configuracion():
    if st.session_state.user["rol"]!="admin": st.error("Solo administradores."); return
    import pandas as pd
    page_header("Configuración del sistema","Estado y mantenimiento de la plataforma")
    reports=load_reports()

    # Estado de persistencia GitHub
    gh = _gh_cfg()
    if gh:
        st.markdown(f"""
        <div style="background:#F0FDF4;border:1px solid #BBF7D0;border-left:4px solid #22C55E;
                    border-radius:0 8px 8px 0;padding:10px 16px;margin-bottom:14px;display:flex;align-items:center;gap:10px">
            <span style="font-size:20px">✅</span>
            <div>
                <div style="font-size:13px;font-weight:700;color:#166534">GitHub conectado — datos persistentes</div>
                <div style="font-size:11px;color:#16a34a">
                    Repositorio: <strong>{gh.get("repo","")}</strong> · Branch: <strong>{gh.get("branch","main")}</strong><br>
                    Los reportes, usuarios y datos MINSAL se guardan automáticamente en GitHub y nunca se pierden.
                </div>
            </div>
        </div>""", unsafe_allow_html=True)
    else:
        st.markdown("""
        <div style="background:#FEF2F2;border:1px solid #FECACA;border-left:4px solid #E24B4A;
                    border-radius:0 8px 8px 0;padding:10px 16px;margin-bottom:14px">
            <div style="font-size:13px;font-weight:700;color:#991B1B">⚠️ GitHub no configurado — datos en riesgo</div>
            <div style="font-size:11px;color:#991B1B;margin-top:4px">
                Sin GitHub configurado, los datos se pierden cuando Streamlit Cloud reinicia la app.<br>
                <strong>Configura st.secrets en Streamlit Cloud para activar la persistencia.</strong>
            </div>
        </div>""", unsafe_allow_html=True)

        with st.expander("📋 Instrucciones para configurar GitHub"):
            st.markdown("""
            **Paso 1 — Crear token en GitHub:**
            1. Ve a github.com → tu foto → **Settings**
            2. **Developer settings** → **Personal access tokens** → **Tokens (classic)**
            3. **Generate new token** → marca **repo** → copia el token (`ghp_...`)

            **Paso 2 — Configurar en Streamlit Cloud:**
            1. Ve a [share.streamlit.io](https://share.streamlit.io) → tu app → **⋮** → **Settings**
            2. Sección **Secrets** → pegar esto:

            ```toml
            [github]
            token = "ghp_TU_TOKEN_AQUI"
            repo = "barolaro/ssmoc_td"
            branch = "main"
            ```

            3. Clic **Save** → la app se reiniciará automáticamente

            **Paso 3 — Crear carpeta data/ en GitHub:**
            En tu repositorio, crea el archivo `data/.gitkeep` para que la carpeta exista.
            """)

    c1,c2=st.columns(2)
    with c1:
        st.subheader("Resumen del sistema")
        st.markdown(f"""| Parámetro | Valor |\n|-----------|-------|\n| Versión | 1.0.0 |\n| Lineamiento | MINSAL v1.0 · Jun 2026 |\n| Meta | ≤ 16% |\n| Establecimientos | {len(ESTABLECIMIENTOS)} |\n| Períodos | {len(PERIODOS)} |\n| Reportes totales | {len(reports)} |\n| Enviados | {len([r for r in reports if r.get("estado")=="enviado"])} |\n| Borradores | {len([r for r in reports if r.get("estado")=="borrador"])} |""")
    with c2:
        st.subheader("Avance por período")
        req=len([e for e in ESTABLECIMIENTOS.values() if e["nivel"] in ["rojo","amarillo"]])
        for p in PERIODOS:
            done=len([r for r in reports if r.get("reporte_id")==p["id"] and r.get("estado")=="enviado"])
            st.markdown(f"**{p['label']}** — {p['periodo']} · Plazo: `{p['fecha_txt']}`")
            st.progress(done/req if req else 0,text=f"{done}/{req} enviados")
    st.divider()
    st.subheader("Datos de referencia SSMOCC (CSV MINSAL)")
    df=pd.DataFrame([{"Establecimiento":e["nombre_corto"],"Nivel":e["nivel"].capitalize(),"% TD 2026":e["pct_2026"],"% TD 2025":e["pct_2025"],"Brecha (pp)":e["brecha"],"Var. (pp)":e["variacion"],"Denominador":e["denominador"],"Numerador":e["numerador"]} for e in ESTABLECIMIENTOS.values()])
    st.dataframe(df,use_container_width=True,hide_index=True)
    st.divider()
    with st.expander("🗑️ Mantenimiento"):
        st.warning("⚠️ Acción irreversible.")
        if st.button("Eliminar TODOS los reportes"):
            save_reports([]); st.success("Reportes eliminados."); st.rerun()


def pg_mis_reportes():
    user = st.session_state.user
    page_header("Ingreso de antecedentes — Anexo N°1",
                "Lineamiento MINSAL v1.0 · Jun 2026 · Subsecretaría de Redes Asistenciales")
    year = st.session_state.get("selected_year", 2026)
    rid_activo = st.session_state.get("selected_report", "R1")
    apply_period_context(year, rid_activo)
    if not periodo_tiene_datos(year, rid_activo):
        mensaje_periodo_sin_datos(year, rid_activo)
        st.info("Los establecimientos no podrán ingresar causas, medidas ni compromisos hasta que exista una carga oficial del período. Esto evita que se informen antecedentes usando indicadores de otro trimestre.")
        return

    if user["rol"] == "admin":
        opciones = {eid: e["nombre_corto"] for eid, e in ESTABLECIMIENTOS.items() if e["nivel"] in ["rojo","amarillo"]}
        eid_sel = st.selectbox("Establecimiento", options=list(opciones.keys()), format_func=lambda x: opciones[x])
    else:
        eid_sel = user.get("establecimiento")
        if not eid_sel: st.error("Sin establecimiento asignado."); return
        if ESTABLECIMIENTOS.get(eid_sel,{}).get("nivel") == "verde":
            est_v = ESTABLECIMIENTOS[eid_sel]
            st.markdown(f"""
            <div style="background:#F0FDF4;border:1px solid #BBF7D0;border-left:5px solid #22C55E;
                        border-radius:8px;padding:24px;text-align:center;margin-top:20px">
                <div style="font-size:42px">✅</div>
                <div style="font-size:17px;font-weight:700;color:#166534;margin-top:8px">Nivel Verde — {est_v['nombre']}</div>
                <div style="font-size:12px;color:#16a34a;margin-top:8px;line-height:1.6">
                    No se requiere remitir antecedentes al MINSAL.<br>
                    El SSMOCC mantendrá el seguimiento interno.
                </div>
            </div>""", unsafe_allow_html=True)
            return

    estab = ESTABLECIMIENTOS.get(eid_sel, {})
    nivel = estab.get("nivel","verde")
    nc = {"rojo":"#E24B4A","amarillo":"#F59E0B","verde":"#22C55E"}[nivel]
    nt = {"rojo":"#991B1B","amarillo":"#92400E","verde":"#166534"}[nivel]
    nb = {"rojo":"#FEF2F2","amarillo":"#FFFBEB","verde":"#F0FDF4"}[nivel]
    nivel_icon = {"rojo":"🔴","amarillo":"🟡","verde":"🟢"}[nivel]
    brecha_s = ("+" if estab.get("brecha",0)>0 else "") + f'{estab.get("brecha",0):.2f} pp'

    st.markdown(f"""
    <div style="background:#1F3864;border-radius:12px 12px 0 0;padding:16px 22px;
                display:flex;justify-content:space-between;align-items:center">
        <div>
            <div style="font-size:16px;font-weight:700;color:white">{estab.get("nombre","")}</div>
            <div style="font-size:11px;color:rgba(255,255,255,.5);margin-top:3px">
                RUT: {estab.get("rut","")} &nbsp;·&nbsp; DEIS: {estab.get("codigo_deis","")}
            </div>
        </div>
        <div style="background:{nb};color:{nt};border-radius:8px;padding:6px 16px;font-size:13px;font-weight:700">
            {nivel_icon} {nivel.upper()}
        </div>
    </div>
    <div style="background:{nb};border:1px solid {nc};border-top:none;border-radius:0 0 12px 12px;
                padding:10px 22px;margin-bottom:20px;display:grid;grid-template-columns:repeat(4,1fr);gap:12px">
        <div style="text-align:center">
            <div style="font-size:10px;color:{nt};text-transform:uppercase">% TD 2026</div>
            <div style="font-size:22px;font-weight:800;color:{nt}">{estab.get("pct_2026",0):.2f}%</div>
        </div>
        <div style="text-align:center">
            <div style="font-size:10px;color:{nt};text-transform:uppercase">% TD 2025</div>
            <div style="font-size:22px;font-weight:700;color:{nt}">{estab.get("pct_2025",0):.2f}%</div>
        </div>
        <div style="text-align:center">
            <div style="font-size:10px;color:{nt};text-transform:uppercase">Brecha vs meta</div>
            <div style="font-size:22px;font-weight:700;color:{nt}">{brecha_s}</div>
        </div>
        <div style="text-align:center">
            <div style="font-size:10px;color:{nt};text-transform:uppercase">Meta 2026</div>
            <div style="font-size:22px;font-weight:700;color:{nt}">16%</div>
        </div>
    </div>""", unsafe_allow_html=True)

    reports_all = load_reports()
    if f"per_{eid_sel}" not in st.session_state:
        st.session_state[f"per_{eid_sel}"] = "R1"

    st.markdown('<div style="font-size:13px;font-weight:600;color:#1F3864;margin-bottom:8px">Seleccionar período</div>', unsafe_allow_html=True)
    cols_per = st.columns(4)
    for i, p in enumerate(PERIODOS):
        r_ex = next((x for x in reports_all if x.get("establecimiento_id")==eid_sel and x.get("reporte_id")==p["id"]), None)
        est_ex = r_ex.get("estado","pendiente") if r_ex else "pendiente"
        icon_p = {"enviado":"✅","borrador":"📝","pendiente":"⬜"}[est_ex]
        sel = st.session_state[f"per_{eid_sel}"] == p["id"]
        border = "3px solid #1F3864" if sel else "1px solid #e2e8f0"
        bg_p = {"enviado":"#F0FDF4","borrador":"#EFF6FF","pendiente":"white"}[est_ex]
        tc_p = {"enviado":"#166534","borrador":"#1E40AF","pendiente":"#64748b"}[est_ex]
        with cols_per[i]:
            st.markdown(f"""
            <div style="background:{bg_p};border:{border};border-radius:10px;padding:12px 8px;text-align:center">
                <div style="font-size:24px">{icon_p}</div>
                <div style="font-size:12px;font-weight:700;color:#1F3864;margin-top:4px">{p["label"]}</div>
                <div style="font-size:10px;color:#64748b">{p["periodo"]}</div>
                <div style="font-size:10px;color:#94a3b8">Plazo: {p["fecha_txt"]}</div>
                <div style="font-size:10px;font-weight:600;color:{tc_p};margin-top:4px">{est_ex.upper()}</div>
            </div>""", unsafe_allow_html=True)
            if st.button(f"Selec. {p['label']}", key=f"selp_{p['id']}_{eid_sel}", use_container_width=True):
                st.session_state[f"per_{eid_sel}"] = p["id"]; st.rerun()

    periodo_id = st.session_state[f"per_{eid_sel}"]
    pinfo = next(p for p in PERIODOS if p["id"] == periodo_id)
    existing = next((r for r in reports_all if r.get("establecimiento_id")==eid_sel and r.get("reporte_id")==periodo_id), None)

    st.markdown("<br>", unsafe_allow_html=True)
    if existing:
        ec = "#166534" if existing.get("estado")=="enviado" else "#1E40AF"
        eb = "#F0FDF4" if existing.get("estado")=="enviado" else "#EFF6FF"
        ei = "✅" if existing.get("estado")=="enviado" else "📝"
        st.markdown(f"""
        <div style="background:{eb};border-radius:8px;padding:10px 16px;margin-bottom:12px;display:flex;align-items:center;gap:10px">
            <span style="font-size:22px">{ei}</span>
            <div>
                <div style="font-size:13px;font-weight:700;color:{ec}">{pinfo["label"]} — {existing.get("estado","borrador").upper()}</div>
                <div style="font-size:11px;color:{ec}">Ingresado por {existing.get("usuario","")} · {existing.get("fecha_ingreso","")[:16]} · Puedes editarlo.</div>
            </div>
        </div>""", unsafe_allow_html=True)

    st.markdown(f'''
    <div style="background:#1F3864;border-radius:8px 8px 0 0;padding:12px 20px">
        <div style="font-size:14px;font-weight:700;color:white">📋 {pinfo["label"]} — {pinfo["periodo"]}</div>
        <div style="font-size:11px;color:rgba(255,255,255,.5);margin-top:2px">Plazo: {pinfo["fecha_txt"]}</div>
    </div>''', unsafe_allow_html=True)

    with st.form(f"frm_{eid_sel}_{periodo_id}", clear_on_submit=False):

        st.markdown('''<div style="background:#EFF6FF;border-left:4px solid #1F3864;border-radius:0 6px 6px 0;padding:10px 16px;margin:12px 0 10px">
            <div style="font-size:13px;font-weight:700;color:#1F3864">1 · Principales causas del resultado observado</div>
            <div style="font-size:11px;color:#3B82F6;margin-top:2px">Seleccione todas las causales aplicables y describa el contexto</div>
        </div>''', unsafe_allow_html=True)

        _dc = [c for c in (existing.get("causas_sel",[]) if existing else []) if c in CAUSALES]
        causas_sel = st.multiselect("Causales de Trato Directo *", CAUSALES, default=_dc)
        causas_desc = st.text_area("Descripción detallada *", value=existing.get("causas_desc","") if existing else "", height=110,
                                   placeholder="Describa las causas específicas del período...")

        st.markdown('<div style="font-size:12px;font-weight:600;color:#374151;margin:10px 0 4px">Datos cuantitativos</div>', unsafe_allow_html=True)
        c1,c2,c3 = st.columns(3)
        monto_td = c1.number_input("💰 Monto TD ($CLP)", min_value=0, step=1_000_000, value=int(existing.get("monto_td",0)) if existing else 0, format="%d")
        n_proc   = c2.number_input("📦 N° procesos TD", min_value=0, step=1, value=int(existing.get("n_proc",0)) if existing else 0)
        pct_per  = c3.number_input("📊 % TD período MINSAL", min_value=0.0, max_value=100.0, step=0.01,
                                   value=float(existing.get("pct_per",estab.get("pct_2026",0.0))) if existing else float(estab.get("pct_2026",0.0)), format="%.2f")

        st.markdown('''<div style="background:#F0FDF4;border-left:4px solid #22C55E;border-radius:0 6px 6px 0;padding:10px 16px;margin:16px 0 10px">
            <div style="font-size:13px;font-weight:700;color:#166534">2 · Medidas implementadas</div>
            <div style="font-size:11px;color:#16A34A;margin-top:2px">Acciones ejecutadas para reducir el uso del Trato Directo</div>
        </div>''', unsafe_allow_html=True)

        med_labels = {"pac":("📅","Actualización Plan Anual de Compras"),"lic":("📄","Inicio procesos licitatorios"),
                      "cm":("🛒","Migración a Convenio Marco"),"cenabast":("🏥","Gestión CENABAST"),
                      "cap":("📚","Capacitación equipo Ley 21.634"),"venc":("⏰","Control vencimiento contratos")}
        ex_med = existing.get("medidas",{}) if existing else {}
        med_sel = {}
        cols_m = st.columns(3)
        for i,(k,(icon_m,lbl)) in enumerate(med_labels.items()):
            med_sel[k] = cols_m[i%3].checkbox(f"{icon_m} {lbl}", value=ex_med.get(k,False), key=f"m_{k}")
        med_desc = st.text_area("Descripción adicional de medidas", value=existing.get("med_desc","") if existing else "", height=80,
                                placeholder="Describa el resultado de las acciones y el impacto observado...")

        st.markdown('''<div style="background:#FFF7ED;border-left:4px solid #F59E0B;border-radius:0 6px 6px 0;padding:10px 16px;margin:16px 0 10px">
            <div style="font-size:13px;font-weight:700;color:#92400E">3 · Compromisos para el próximo período</div>
            <div style="font-size:11px;color:#D97706;margin-top:2px">Acciones concretas con plazos verificables</div>
        </div>''', unsafe_allow_html=True)

        compromisos = st.text_area("Compromisos adoptados *", value=existing.get("compromisos","") if existing else "", height=110,
                                   placeholder="1. Iniciar licitación para [insumo] antes del [fecha]")
        c4,c5 = st.columns(2)
        meta_prox  = c4.number_input("🎯 Meta % TD próximo período", min_value=0.0, max_value=100.0, step=0.5,
                                     value=float(existing.get("meta_prox",16.0)) if existing else 16.0, format="%.1f")
        fecha_comp = c5.date_input("📆 Fecha comprometida",
                                   value=datetime.date.fromisoformat(existing["fecha_comp"])
                                   if existing and existing.get("fecha_comp") else datetime.date(2026,8,31))

        st.markdown('''<div style="background:#F8FAFC;border-left:4px solid #64748B;border-radius:0 6px 6px 0;padding:10px 16px;margin:16px 0 10px">
            <div style="font-size:13px;font-weight:700;color:#374151">4 · Responsable del reporte</div>
        </div>''', unsafe_allow_html=True)

        c6,c7,c8 = st.columns(3)
        resp_nombre = c6.text_input("👤 Nombre completo", value=existing.get("resp_nombre",user["nombre"]) if existing else user["nombre"])
        resp_cargo  = c7.text_input("💼 Cargo", value=existing.get("resp_cargo","Jefe/a de Abastecimiento") if existing else "Jefe/a de Abastecimiento")
        resp_email  = c8.text_input("📧 Correo", value=existing.get("resp_email",user.get("email","")) if existing else user.get("email",""))
        obs = st.text_area("💬 Observaciones (opcional)", value=existing.get("obs","") if existing else "", height=60)

        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown('''<div style="background:#F8FAFC;border:1px solid #E2E8F0;border-radius:8px;padding:10px 14px;margin-bottom:10px;font-size:11px;color:#64748b">
            <strong>* Obligatorios para enviar.</strong> Puedes guardar como borrador y completar después.
        </div>''', unsafe_allow_html=True)

        cg,ce,_ = st.columns([1,1,2])
        guardar = cg.form_submit_button("💾  Guardar borrador", use_container_width=True)
        enviar  = ce.form_submit_button("📤  Enviar a SSMOCC",  use_container_width=True, type="primary")

        if guardar or enviar:
            ok = True
            if enviar:
                errs = []
                if not causas_sel:          errs.append("⚠️ Seleccione al menos una causal.")
                if not causas_desc.strip(): errs.append("⚠️ Complete la descripción de causas.")
                if not compromisos.strip(): errs.append("⚠️ Complete los compromisos.")
                if not resp_nombre.strip(): errs.append("⚠️ Ingrese el nombre del responsable.")
                if errs:
                    for e in errs: st.error(e)
                    ok = False
            if ok:
                data = {"establecimiento_id":eid_sel,"establecimiento_nombre":estab["nombre"],"reporte_id":periodo_id,"periodo":pinfo["periodo"],"periodo_label":pinfo["label"],"nivel_riesgo":nivel,"pct_2026":estab.get("pct_2026"),"pct_2025":estab.get("pct_2025"),"pct_per":pct_per,"monto_td":monto_td,"n_proc":n_proc,"causas_sel":causas_sel,"causas_desc":causas_desc,"medidas":med_sel,"med_desc":med_desc,"compromisos":compromisos,"meta_prox":meta_prox,"fecha_comp":str(fecha_comp),"resp_nombre":resp_nombre,"resp_cargo":resp_cargo,"resp_email":resp_email,"obs":obs,"estado":"enviado" if enviar else "borrador","usuario":user["username"],"fecha_ingreso":str(datetime.datetime.now().isoformat())}
                upsert_report(data)
                if enviar: st.success(f"✅ **Reporte {pinfo['label']}** enviado exitosamente."); st.balloons()
                else: st.info("💾 Borrador guardado correctamente.")
                st.rerun()


def pg_actualizar_datos():
    """Página para cargar datos oficiales MINSAL por Año + Período, con trazabilidad histórica."""
    if st.session_state.user["rol"] != "admin":
        st.error("Solo administradores."); return
    page_header("Carga Oficial MINSAL", "Carga histórica por año y período · dashboard, boletín y Excel se actualizan automáticamente")
    st.markdown("""<div style="background:#EFF6FF;border:1px solid #BFDBFE;border-left:4px solid #1F3864;border-radius:0 8px 8px 0;padding:12px 16px;margin-bottom:16px"><div style="font-size:13px;font-weight:700;color:#1F3864;margin-bottom:6px">📋 Flujo institucional de actualización</div><div style="font-size:12px;color:#1E40AF;line-height:1.7">Selecciona <strong>año + período</strong>, sube el CSV oficial MINSAL y confirma la carga.<br>La información queda archivada históricamente y puede reutilizarse para dashboard, boletín ejecutivo y Excel MINSAL.<br><strong>Los reportes ingresados por los establecimientos se mantienen separados por período.</strong></div></div>""", unsafe_allow_html=True)
    c1, c2 = st.columns([1, 2])
    with c1:
        year = st.selectbox("Año de reporte", list(range(2026, 2031)), index=0)
    with c2:
        reporte_id = st.selectbox("Período", [p["id"] for p in PERIODOS], format_func=lambda x: f"{get_periodo_info(x)['label']} — {get_periodo_info(x)['periodo']} · entrega {get_periodo_info(x)['fecha_txt']}")
    st.subheader(f"Estado actual — {periodo_display(year, reporte_id)}")
    apply_period_context(year, reporte_id)
    import pandas as pd
    if not periodo_tiene_datos(year, reporte_id):
        mensaje_periodo_sin_datos(year, reporte_id, detalle=False)
    else:
        rows_act = []
        for eid, e in ESTABLECIMIENTOS.items():
            rows_act.append({"Establecimiento": e["nombre_corto"], "Nivel": e["nivel"].capitalize(), "% TD 2026": f"{e['pct_2026']:.2f}%", "% TD 2025": f"{e['pct_2025']:.2f}%", "Brecha": f"{e['brecha']:+.2f} pp", "Denominador ($)": f"${e['denominador']:,.0f}", "Numerador ($)": f"${e['numerador']:,.0f}", "Última actualización": e.get("ultima_actualizacion", "Sin carga histórica")})
        st.dataframe(pd.DataFrame(rows_act), use_container_width=True, hide_index=True)
    st.markdown("---")
    st.subheader("📤 Subir CSV oficial MINSAL")
    st.markdown("""<div style="background:#FFFBEB;border:1px solid #FDE68A;border-radius:8px;padding:12px 14px;margin-bottom:12px;font-size:12px;color:#92400E"><strong>Formato esperado:</strong> CSV con separador <code>;</code> y columnas: <code>Codigo DEIS · RUT · Establecimiento · Servicio de Salud · Denominador · Numerador · % Trato Directo 2026 · % Trato Directo periodo equivalente 2025 · Brecha vs meta · Variacion vs 2025 · Nivel de riesgo</code>.</div>""", unsafe_allow_html=True)
    uploaded = st.file_uploader("Seleccionar archivo CSV del MINSAL", type=["csv"])
    if uploaded:
        try:
            file_bytes = uploaded.read()
            updates, errors, rows_found = parse_csv_minsal(file_bytes)
            st.success(f"Archivo leído correctamente: {rows_found} establecimiento(s) reconocidos.")
            if errors:
                with st.expander(f"⚠️ {len(errors)} advertencia(s)"):
                    for e in errors: st.warning(e)
            if updates:
                preview_rows = []
                for eid, vals in updates.items():
                    e_old = ESTABLECIMIENTOS.get(eid, {})
                    preview_rows.append({"Establecimiento": e_old.get("nombre_corto", eid), "% TD nuevo": f"{vals['pct_2026']:.2f}%", "% TD anterior período activo": f"{e_old.get('pct_2026',0):.2f}%", "Variación": f"{vals['pct_2026']-e_old.get('pct_2026',0):+.2f} pp", "Nivel nuevo": vals["nivel"].capitalize(), "Nivel anterior": e_old.get("nivel","").capitalize()})
                st.markdown("**Vista previa antes de guardar:**")
                st.dataframe(pd.DataFrame(preview_rows), use_container_width=True, hide_index=True)
                col1, col2 = st.columns([1, 3])
                with col1:
                    if st.button("✅ Confirmar carga histórica", type="primary", use_container_width=True):
                        save_datos_periodo(year, reporte_id, updates, uploaded.name)
                        save_datos_minsal(updates)
                        st.session_state.selected_year = year
                        st.session_state.selected_report = reporte_id
                        apply_period_context(year, reporte_id)
                        st.success(f"Datos guardados para {periodo_display(year, reporte_id)}.")
                        st.info("Ya puedes revisar el dashboard, generar el boletín y exportar el Excel MINSAL.")
                        st.rerun()
                with col2:
                    st.markdown(f"<div style='padding:8px 0;font-size:12px;color:#64748b'>Esta carga quedará archivada como <b>{periodo_display(year, reporte_id)}</b>. No sobrescribe otros períodos.</div>", unsafe_allow_html=True)
            else:
                st.error("No se reconoció ningún establecimiento SSMOCC en el CSV.")
        except Exception as ex:
            st.error(f"Error al leer el archivo: {ex}")
    st.markdown("---")
    data = load_datos_periodos()
    st.subheader("📚 Cargas históricas registradas")
    hist = []
    for y, periods in sorted(data.items()):
        for rid, obj in periods.items():
            meta = obj.get("metadata", {})
            hist.append({"Año": y, "Período": f"{rid} — {get_periodo_info(rid)['periodo']}", "Establecimientos": meta.get("establecimientos", len(obj.get("establecimientos", {}))), "Fuente": meta.get("fuente", "—"), "Fecha carga": meta.get("fecha_carga", "—")})
    if hist:
        st.dataframe(pd.DataFrame(hist), use_container_width=True, hide_index=True)
    else:
        st.info("Aún no existen cargas históricas. La primera carga creará la base del período.")


def _resumen_establecimientos_df():
    import pandas as pd
    return pd.DataFrame([{"Establecimiento": e["nombre_corto"], "Nivel": e["nivel"].capitalize(), "% TD 2026": e["pct_2026"], "% TD 2025": e["pct_2025"], "Brecha pp": e["brecha"], "Variación pp": e["variacion"], "Numerador": e["numerador"], "Denominador": e["denominador"]} for e in ESTABLECIMIENTOS.values()]).sort_values("% TD 2026", ascending=False)


def generar_boletin_html(year:int, reporte_id:str):
    apply_period_context(year, reporte_id)
    pinfo = get_periodo_info(reporte_id)
    total_num = sum(e["numerador"] for e in ESTABLECIMIENTOS.values())
    total_den = sum(e["denominador"] for e in ESTABLECIMIENTOS.values())
    pct = round(total_num/total_den*100, 2) if total_den else 0
    rojos = [e for e in ESTABLECIMIENTOS.values() if e["nivel"] == "rojo"]
    amarillos = [e for e in ESTABLECIMIENTOS.values() if e["nivel"] == "amarillo"]
    verdes = [e for e in ESTABLECIMIENTOS.values() if e["nivel"] == "verde"]
    rows = "".join([f"<tr><td>{e['nombre_corto']}</td><td>{e['pct_2026']:.2f}%</td><td>{e['pct_2025']:.2f}%</td><td>{e['brecha']:+.2f}</td><td><b>{e['nivel'].capitalize()}</b></td></tr>" for e in sorted(ESTABLECIMIENTOS.values(), key=lambda x: -x['pct_2026'])])
    fecha = datetime.datetime.now().strftime("%d/%m/%Y %H:%M")
    return f"""<!doctype html><html><head><meta charset='utf-8'><title>Boletín TD SSMOCC</title><style>body{{font-family:Arial,sans-serif;margin:0;background:#f8fafc;color:#0f172a}}.cover{{background:#1F3864;color:white;padding:34px 46px;border-bottom:8px solid #C0392B}}.cover h1{{margin:0;font-size:30px}}.wrap{{padding:26px 46px}}.kpis{{display:grid;grid-template-columns:repeat(4,1fr);gap:12px}}.card{{background:white;border:1px solid #e2e8f0;border-top:4px solid #1F3864;border-radius:10px;padding:16px}}.lbl{{font-size:11px;color:#64748b;text-transform:uppercase}}.val{{font-size:26px;font-weight:800;color:#1F3864;margin-top:6px}}table{{width:100%;border-collapse:collapse;background:white;margin-top:18px}}th{{background:#1F3864;color:white;text-align:left;padding:10px;font-size:12px}}td{{border-bottom:1px solid #e2e8f0;padding:9px;font-size:12px}}.box{{background:white;border-left:5px solid #C0392B;padding:14px;margin:18px 0;border-radius:6px}}</style></head><body><div class='cover'><h1>Boletín Ejecutivo · Indicador Trato Directo</h1><p>Servicio de Salud Metropolitano Occidente · {year} · {pinfo['periodo']} · plazo MINSAL {pinfo['fecha_txt']}</p></div><div class='wrap'><div class='kpis'><div class='card'><div class='lbl'>% TD SSMOCC</div><div class='val'>{pct:.2f}%</div></div><div class='card'><div class='lbl'>Numerador</div><div class='val'>${total_num/1e9:.1f} MM</div></div><div class='card'><div class='lbl'>Denominador</div><div class='val'>${total_den/1e9:.1f} MM</div></div><div class='card'><div class='lbl'>Semáforo</div><div class='val'>{len(rojos)} R · {len(amarillos)} A · {len(verdes)} V</div></div></div><div class='box'><b>Resumen ejecutivo:</b> El indicador consolidado del Servicio alcanza <b>{pct:.2f}%</b>, con meta institucional de referencia ≤16%. Los establecimientos en nivel rojo y amarillo requieren análisis de causas, medidas correctivas y compromisos para el siguiente período.</div><h2>Ranking por establecimiento</h2><table><thead><tr><th>Establecimiento</th><th>% TD 2026</th><th>% TD 2025</th><th>Brecha pp</th><th>Nivel</th></tr></thead><tbody>{rows}</tbody></table><h2>Recomendaciones de gestión</h2><ul><li>Priorizar revisión de establecimientos en rojo.</li><li>Migrar compras recurrentes a mecanismos competitivos o convenios disponibles.</li><li>Controlar vencimiento de contratos y plan anual de compras.</li><li>Revisar compromisos pendientes antes de la siguiente fecha de reporte.</li></ul><p style='font-size:11px;color:#64748b'>Generado automáticamente por Monitor TD SSMOCC · {fecha}</p></div></body></html>"""


def pg_boletines():
    if st.session_state.user["rol"] != "admin": st.error("Solo administradores."); return
    import pandas as pd
    page_header("Boletines Ejecutivos", "Generación automática del boletín general del dashboard")
    year = st.session_state.get("selected_year", 2026)
    rid = st.session_state.get("selected_report", "R1")
    apply_period_context(year, rid)
    st.markdown(f"**Período activo:** {periodo_display(year, rid)}")
    if not periodo_tiene_datos(year, rid):
        mensaje_periodo_sin_datos(year, rid)
        st.stop()
    df = _resumen_establecimientos_df()
    st.dataframe(df, use_container_width=True, hide_index=True)
    html = generar_boletin_html(year, rid)
    fecha = datetime.datetime.now().strftime("%Y%m%d_%H%M")
    st.download_button("⬇️ Descargar boletín ejecutivo HTML", html.encode("utf-8"), f"Boletin_TD_SSMOCC_{year}_{rid}_{fecha}.html", "text/html", type="primary", use_container_width=True)
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        df.to_excel(w, sheet_name="Resumen dashboard", index=False)
    buf.seek(0)
    st.download_button("⬇️ Descargar respaldo Excel del boletín", buf, f"Boletin_TD_SSMOCC_{year}_{rid}_{fecha}.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)


def pg_historico():
    if st.session_state.user["rol"] != "admin": st.error("Solo administradores."); return
    import pandas as pd
    page_header("Histórico y Comparativos", "Evolución por año y período")
    data = load_datos_periodos()
    rows = []
    for y, periods in sorted(data.items()):
        for rid, obj in sorted(periods.items()):
            ests = obj.get("establecimientos", {})
            den = sum(v.get("denominador",0) for v in ests.values())
            num = sum(v.get("numerador",0) for v in ests.values())
            rows.append({"Año": y, "Período": f"{rid} — {get_periodo_info(rid)['periodo']}", "% TD SSMOCC": round(num/den*100,2) if den else 0, "Rojo": sum(1 for v in ests.values() if v.get("nivel")=="rojo"), "Amarillo": sum(1 for v in ests.values() if v.get("nivel")=="amarillo"), "Verde": sum(1 for v in ests.values() if v.get("nivel")=="verde"), "Fecha carga": obj.get("metadata",{}).get("fecha_carga", "—")})
    if not rows:
        st.info("Aún no existen períodos cargados históricamente."); return
    df = pd.DataFrame(rows)
    st.dataframe(df, use_container_width=True, hide_index=True)
    try:
        import plotly.express as px
        fig = px.line(df, x="Período", y="% TD SSMOCC", markers=True, title="Evolución % TD SSMOCC")
        fig.add_hline(y=16, line_dash="dash", line_color="#1F3864", annotation_text="Meta 16%")
        st.plotly_chart(fig, use_container_width=True)
    except Exception:
        pass



# ═══════════════════════════════════════════════════════════════════════
# PÁGINA EXPORTAR — ANEXO N°1 MINSAL OFICIAL
# Esta definición reemplaza la anterior y deja el Excel alineado al formato
# solicitado: solo establecimientos ROJO/AMARILLO que hayan entregado reporte.
# ═══════════════════════════════════════════════════════════════════════
def pg_exportar():
    if st.session_state.user["rol"] != "admin":
        st.error("Solo administradores."); return
    import pandas as pd
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    page_header("Exportar Anexo N°1 MINSAL", "Formato oficial de entrega de antecedentes por Servicio de Salud")
    reports = load_reports()
    year = st.session_state.get("selected_year", 2026)
    ps = st.selectbox(
        "Período",
        options=[p["id"] for p in PERIODOS],
        index=[p["id"] for p in PERIODOS].index(st.session_state.get("selected_report", "R1")) if st.session_state.get("selected_report", "R1") in [p["id"] for p in PERIODOS] else 0,
        format_func=lambda x: next(f"{p['label']} — {p['periodo']} (plazo: {p['fecha_txt']})" for p in PERIODOS if p["id"] == x)
    )
    pinfo = next(p for p in PERIODOS if p["id"] == ps)
    st.session_state.selected_report = ps
    st.session_state.selected_year = year
    apply_period_context(year, ps)
    if not periodo_tiene_datos(year, ps):
        mensaje_periodo_sin_datos(year, ps)
        st.info("El Anexo N°1 MINSAL queda bloqueado hasta cargar la base oficial del período. Así se evita exportar información del período anterior.")
        return

    # Solo deben reportar los establecimientos categorizados en rojo o amarillo.
    obligatorios = {eid: e for eid, e in ESTABLECIMIENTOS.items() if e.get("nivel") in ["rojo", "amarillo"]}
    enviados = []
    borradores = []
    pendientes = []

    for eid, e in obligatorios.items():
        r = next((x for x in reports if x.get("establecimiento_id") == eid and x.get("reporte_id") == ps and x.get("estado") == "enviado"), None)
        b = next((x for x in reports if x.get("establecimiento_id") == eid and x.get("reporte_id") == ps and x.get("estado") == "borrador"), None)
        if r:
            enviados.append(r)
        elif b:
            borradores.append(b)
        else:
            pendientes.append((eid, e))

    c1, c2, c3 = st.columns(3)
    c1.metric("Obligatorios rojo/amarillo", len(obligatorios))
    c2.metric("Reportes enviados", len(enviados))
    c3.metric("Pendientes", len(pendientes))

    st.markdown(f"""
    <div style="background:#EFF6FF;border:1px solid #BFDBFE;border-left:4px solid #1F3864;border-radius:0 8px 8px 0;padding:12px 16px;margin:10px 0 16px;font-size:12px;color:#1E40AF;line-height:1.6">
        Este módulo genera el <strong>Anexo N°1 MINSAL</strong> usando exclusivamente los antecedentes entregados por los establecimientos en categoría <strong>roja y amarilla</strong> para <strong>{year} · {pinfo['periodo']}</strong>.<br>
        Los establecimientos verdes quedan fuera del Anexo, salvo que el lineamiento solicite lo contrario.
    </div>
    """, unsafe_allow_html=True)

    def _texto_causas(r):
        causas = "; ".join(r.get("causas_sel", []))
        desc = r.get("causas_desc", "")
        return (causas + (" | " if causas and desc else "") + desc).strip()

    def _texto_medidas(r):
        med = r.get("medidas", {})
        med_labels = {
            "pac": "Actualización Plan Anual de Compras",
            "lic": "Inicio procesos licitatorios",
            "cm": "Migración a Convenio Marco",
            "cenabast": "Gestión CENABAST",
            "cap": "Capacitación equipo Ley 21.634",
            "venc": "Control vencimiento contratos",
        }
        txt = "; ".join(lbl for k, lbl in med_labels.items() if med.get(k))
        desc = r.get("med_desc", "")
        return (txt + (" | " if txt and desc else "") + desc).strip()

    rows = []
    for r in sorted(enviados, key=lambda x: (x.get("nivel_riesgo", ""), x.get("establecimiento_nombre", ""))):
        rows.append({
            "Servicio de salud": "Servicio de Salud Metropolitano Occidente",
            "Establecimiento": r.get("establecimiento_nombre", ""),
            "Nivel de Riesgo": r.get("nivel_riesgo", "").capitalize(),
            "Período informado": f"{year} · {r.get('periodo', pinfo['periodo'])}",
            "Principales causas": _texto_causas(r),
            "Medidas implementadas": _texto_medidas(r),
            "Compromisos": r.get("compromisos", ""),
            "Responsable": (r.get("resp_nombre", "") + " - " + r.get("resp_cargo", "")).strip(" -"),
            "Fecha comprometida": r.get("fecha_comp", ""),
        })

    cols_minsal = ["Servicio de salud", "Establecimiento", "Nivel de Riesgo", "Período informado", "Principales causas", "Medidas implementadas", "Compromisos", "Responsable", "Fecha comprometida"]

    if rows:
        df = pd.DataFrame(rows, columns=cols_minsal)
        st.markdown("**Vista previa del Anexo N°1 MINSAL**")
        st.dataframe(df, use_container_width=True, hide_index=True)
    else:
        df = pd.DataFrame(columns=cols_minsal)
        st.warning("Aún no existen reportes enviados para establecimientos rojos/amarillos en este período.")

    if borradores:
        with st.expander(f"📝 {len(borradores)} reporte(s) en borrador, no incluidos en el Excel oficial"):
            for r in borradores:
                st.markdown(f"- **{r.get('establecimiento_nombre','')}** — {r.get('nivel_riesgo','').capitalize()}")
    if pendientes:
        with st.expander(f"⚠️ {len(pendientes)} establecimiento(s) pendiente(s) de enviar"):
            for eid, e in pendientes:
                st.markdown(f"- **{e['nombre']}** — {e['nivel'].capitalize()} ({e['pct_2026']:.2f}%)")

    # Construcción Excel con formato similar al Anexo del lineamiento.
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        # Hoja oficial exacta
        df.to_excel(w, sheet_name="Anexo N1 MINSAL", startrow=5, index=False)
        ws = w.book["Anexo N1 MINSAL"]
        ws.merge_cells("A1:I1")
        ws["A1"] = "6.    Anexo N° 1. Formato de entrega de antecedentes por Servicio de Salud"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")
        ws.merge_cells("A3:I4")
        ws["A3"] = "Los antecedentes deberán ser remitidos utilizando el siguiente formato, con el propósito de estandarizar la información consolidada enviada por cada Servicio de Salud."
        ws["A3"].alignment = Alignment(wrap_text=True, vertical="center")
        ws["A3"].font = Font(size=12)

        header_fill = PatternFill("solid", fgColor="F3F4F6")
        thin = Side(style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for cell in ws[6]:
            cell.font = Font(bold=True, size=11)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        for row in ws.iter_rows(min_row=7, max_row=ws.max_row, min_col=1, max_col=9):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = border
        widths = [20, 34, 16, 18, 38, 38, 38, 30, 20]
        for i, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width
        ws.row_dimensions[1].height = 28
        ws.row_dimensions[3].height = 46
        ws.row_dimensions[6].height = 48
        for r in range(7, max(ws.max_row + 1, 8)):
            ws.row_dimensions[r].height = 72
        ws.freeze_panes = "A7"

        # Hoja de control interno SSMOCC
        control_rows = []
        for eid, e in obligatorios.items():
            estado = "Enviado" if any(r.get("establecimiento_id") == eid for r in enviados) else "Borrador" if any(r.get("establecimiento_id") == eid for r in borradores) else "Pendiente"
            control_rows.append({
                "Establecimiento": e["nombre"],
                "Nivel": e["nivel"].capitalize(),
                "% TD 2026": e.get("pct_2026", 0),
                "Estado reporte": estado,
                "Incluido Anexo": "Sí" if estado == "Enviado" else "No",
            })
        pd.DataFrame(control_rows).to_excel(w, sheet_name="Control interno", index=False)

    buf.seek(0)
    fecha = datetime.datetime.now().strftime("%Y%m%d_%H%M")
    st.download_button(
        "⬇️ Descargar Excel Anexo N°1 MINSAL",
        buf,
        f"SSMOCC_AnexoN1_MINSAL_{year}_{ps}_{fecha}.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        type="primary",
        disabled=df.empty,
    )

    csv_bytes = df.to_csv(index=False, encoding="utf-8-sig").encode("utf-8-sig")
    st.download_button(
        "⬇️ Descargar CSV Anexo N°1",
        csv_bytes,
        f"SSMOCC_AnexoN1_MINSAL_{year}_{ps}_{fecha}.csv",
        "text/csv",
        use_container_width=True,
        disabled=df.empty,
    )

# ═══════════════════════════════════════════════════════════════════════
# ROUTING
# ═══════════════════════════════════════════════════════════════════════
if st.session_state.user is None:
    page_login()
else:
    render_topbar()
    pg = st.session_state.page
    if   pg == "dashboard":      pg_dashboard()
    elif pg == "mis_reportes":   pg_mis_reportes()
    elif pg == "todos_reportes": pg_todos_reportes()
    elif pg == "exportar":       pg_exportar()
    elif pg == "boletines":      pg_boletines()
    elif pg == "historico":      pg_historico()
    elif pg == "usuarios":       pg_usuarios()
    elif pg == "configuracion":   pg_configuracion()
    elif pg == "actualizar_datos": pg_actualizar_datos()
    else:                         pg_dashboard()
